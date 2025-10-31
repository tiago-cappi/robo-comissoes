"""
Adapter FastAPI para orquestração do robô de comissões.

Este adapter apenas orquestra chamadas ao robô Python existente,
sem alterar a lógica de negócio.
"""

import os
import json
import subprocess
import uuid
import asyncio
import sys
from pathlib import Path
from typing import Optional, List, Dict, Any
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
import aiofiles
from dotenv import load_dotenv

# Carregar variáveis de ambiente
load_dotenv()

# Configuração
ROBO_ROOT_PATH = os.getenv("ROBO_ROOT_PATH", os.getcwd())
PROGRESS_FILE = os.path.join(ROBO_ROOT_PATH, "progress.json")

app = FastAPI(
    title="Adapter Robô de Comissões",
    description="Backend adapter para orquestração do robô de comissões",
    version="1.0.0"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== MODELS ====================

class BulkApplyRequest(BaseModel):
    escopo: Dict[str, List[str]]  # ex: {"linha": ["A", "B"], "grupo": ["G1"]}
    campos: Dict[str, Any]  # campos a definir
    modo: str  # "criar" ou "atualizar"
    previewOnly: bool = True


class SaveRequest(BaseModel):
    data: List[Dict[str, Any]]
    preserve_columns: bool = True


class ProgressResponse(BaseModel):
    job_id: str
    percent: float
    etapa: str
    mensagens: List[str]
    status: str  # "em_andamento", "concluido", "erro"


# ==================== HELPER FUNCTIONS ====================

def get_regras_path() -> Path:
    """Retorna caminho do arquivo Regras_Comissoes.xlsx"""
    return Path(ROBO_ROOT_PATH) / "Regras_Comissoes.xlsx"


def get_resultado_path() -> Optional[Path]:
    """Retorna caminho do arquivo de resultado mais recente"""
    pattern = "Comissoes_Calculadas_*.xlsx"
    files = list(Path(ROBO_ROOT_PATH).glob(pattern))
    if not files:
        return None
    # Ordenar por data de modificação (mais recente primeiro)
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0]


def read_excel_sheet(filepath: Path, sheet_name: str) -> pd.DataFrame:
    """Lê uma aba do Excel preservando ordem de colunas"""
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str, keep_default_na=False)
        # Preservar ordem original das colunas
        return df
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler aba {sheet_name}: {str(e)}")


def write_excel_sheet(filepath: Path, sheet_name: str, df: pd.DataFrame, preserve_order: bool = True):
    """Escreve uma aba no Excel preservando ordem de colunas"""
    try:
        # Se arquivo existe, carregar para preservar outras abas
        if filepath.exists():
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # Criar novo arquivo
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar aba {sheet_name}: {str(e)}")


# ==================== ENDPOINTS - REGRAS ====================

@app.get("/regras/abas")
async def listar_abas_regras():
    """Lista todas as abas do arquivo Regras_Comissoes.xlsx"""
    regras_path = get_regras_path()
    if not regras_path.exists():
        return {"abas": []}
    
    try:
        wb = load_workbook(regras_path, read_only=True)
        return {"abas": wb.sheetnames}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler arquivo de regras: {str(e)}")


@app.get("/regras/aba/{nome_aba}")
async def ler_aba_regras(
    nome_aba: str,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=1000),
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = Query("asc", regex="^(asc|desc)$"),
    filters: Optional[str] = None,  # JSON string com filtros
):
    """Lê uma aba do Regras_Comissoes.xlsx com paginação e filtros"""
    regras_path = get_regras_path()
    if not regras_path.exists():
        raise HTTPException(status_code=404, detail="Arquivo Regras_Comissoes.xlsx não encontrado")
    
    df = read_excel_sheet(regras_path, nome_aba)
    
    # Aplicar filtros
    if filters:
        try:
            filter_dict = json.loads(filters)
            for col, value in filter_dict.items():
                if col in df.columns and value:
                    df = df[df[col].astype(str).str.contains(str(value), case=False, na=False)]
        except Exception:
            pass
    
    # Ordenação
    if sort_by and sort_by in df.columns:
        ascending = sort_order == "asc"
        df = df.sort_values(by=sort_by, ascending=ascending)
    
    # Paginação
    total = len(df)
    start = (page - 1) * size
    end = start + size
    df_page = df.iloc[start:end]
    
    return {
        "data": df_page.to_dict(orient="records"),
        "total": total,
        "page": page,
        "size": size,
        "columns": list(df.columns)
    }


@app.post("/regras/aba/{nome_aba}/save")
async def salvar_aba_regras(nome_aba: str, request: SaveRequest):
    """Salva alterações em uma aba do Regras_Comissoes.xlsx"""
    regras_path = get_regras_path()
    
    # Ler aba atual para preservar colunas
    if regras_path.exists():
        df_existing = read_excel_sheet(regras_path, nome_aba)
        columns_order = list(df_existing.columns)
    else:
        columns_order = list(request.data[0].keys()) if request.data else []
    
    # Criar DataFrame
    df_new = pd.DataFrame(request.data)
    
    # Preservar ordem de colunas
    if request.preserve_columns and columns_order:
        # Adicionar colunas faltantes
        for col in columns_order:
            if col not in df_new.columns:
                df_new[col] = ""
        # Reordenar
        df_new = df_new[columns_order]
    
    # Salvar
    write_excel_sheet(regras_path, nome_aba, df_new, preserve_order=True)
    
    return {"success": True, "message": f"Aba {nome_aba} salva com sucesso"}


@app.post("/regras/aba/{nome_aba}/apply-bulk")
async def aplicar_massa_regras(nome_aba: str, request: BulkApplyRequest):
    """Aplica alterações em massa na aba"""
    regras_path = get_regras_path()
    if not regras_path.exists():
        raise HTTPException(status_code=404, detail="Arquivo Regras_Comissoes.xlsx não encontrado")
    
    df = read_excel_sheet(regras_path, nome_aba)
    
    # Aplicar filtros de escopo
    mask = pd.Series([True] * len(df))
    for col, values in request.escopo.items():
        if col in df.columns:
            mask = mask & df[col].isin(values)
    
    df_filtered = df[mask].copy()
    
    if request.previewOnly:
        # Pré-visualização
        if request.modo == "criar":
            # Gerar combinações
            # Implementação simplificada - criar novas linhas
            preview_data = []
            for idx, row in df_filtered.iterrows():
                new_row = row.to_dict()
                for key, value in request.campos.items():
                    new_row[key] = value
                preview_data.append(new_row)
        else:
            # Atualizar existentes
            preview_data = df_filtered.copy()
            for key, value in request.campos.items():
                preview_data[key] = value
            preview_data = preview_data.to_dict(orient="records")
        
        return {
            "preview": preview_data[:100],  # Limitar preview
            "total_afetadas": len(df_filtered),
            "previewOnly": True
        }
    else:
        # Aplicar alterações
        if request.modo == "criar":
            # Criar novas linhas
            new_rows = []
            for idx, row in df_filtered.iterrows():
                new_row = row.to_dict()
                for key, value in request.campos.items():
                    new_row[key] = value
                new_rows.append(new_row)
            df_new = pd.DataFrame(new_rows)
            df = pd.concat([df, df_new], ignore_index=True)
        else:
            # Atualizar existentes
            for key, value in request.campos.items():
                if key in df.columns:
                    df.loc[mask, key] = value
        
        write_excel_sheet(regras_path, nome_aba, df, preserve_order=True)
        return {"success": True, "total_afetadas": len(df_filtered)}


# ==================== ENDPOINTS - UPLOADS ====================

@app.post("/upload/analise")
async def upload_analise(file: UploadFile = File(...)):
    """Upload do arquivo Analise_Comercial_Completa"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Nome de arquivo inválido")
    
    # Determinar extensão
    ext = Path(file.filename).suffix.lower()
    if ext not in ['.xlsx', '.csv']:
        raise HTTPException(status_code=400, detail="Formato inválido. Use .xlsx ou .csv")
    
    # Salvar na raiz do projeto
    filename = f"Analise_Comercial_Completa{ext}"
    filepath = Path(ROBO_ROOT_PATH) / filename
    
    async with aiofiles.open(filepath, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    return {"success": True, "filename": filename, "message": "Arquivo salvo com sucesso"}


@app.post("/upload/fin_adcli")
async def upload_fin_adcli(file: UploadFile = File(...)):
    """Upload do arquivo fin_adcli_pg_m3.xls"""
    if not file.filename.endswith('.xls'):
        raise HTTPException(status_code=400, detail="Formato inválido. Use .xls")
    
    filepath = Path(ROBO_ROOT_PATH) / "fin_adcli_pg_m3.xls"
    async with aiofiles.open(filepath, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    return {"success": True, "filename": "fin_adcli_pg_m3.xls"}


@app.post("/upload/fin_conci")
async def upload_fin_conci(file: UploadFile = File(...)):
    """Upload do arquivo fin_conci_adcli_m3.xls"""
    if not file.filename.endswith('.xls'):
        raise HTTPException(status_code=400, detail="Formato inválido. Use .xls")
    
    filepath = Path(ROBO_ROOT_PATH) / "fin_conci_adcli_m3.xls"
    async with aiofiles.open(filepath, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    return {"success": True, "filename": "fin_conci_adcli_m3.xls"}


@app.post("/upload/analise_financeira")
async def upload_analise_financeira(file: UploadFile = File(...)):
    """Upload do arquivo Análise Financeira.xlsx"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Formato inválido. Use .xlsx")
    
    filepath = Path(ROBO_ROOT_PATH) / "Análise Financeira.xlsx"
    async with aiofiles.open(filepath, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    return {"success": True, "filename": "Análise Financeira.xlsx"}


# ==================== ENDPOINTS - EXECUÇÃO ====================

# Dicionário para armazenar processos ativos
processos_ativos: Dict[str, subprocess.Popen] = {}

# Fases esperadas do cálculo (para estimar progresso)
FASES_CALCULO = [
    ("Iniciando...", 0),
    ("Carregando arquivos...", 10),
    ("Validando dados...", 15),
    ("Pré-processando informações...", 20),
    ("Calculando valores realizados agregados...", 30),
    ("Calculando comissões e FC item a item...", 50),
    ("Carregando estado de recebimentos e aplicando adiantamentos...", 70),
    ("Executando reconciliações de processos quitados...", 85),
    ("Gerando arquivos de saída...", 95),
    ("Concluído", 100),
]

async def monitorar_processo(job_id: str, process: subprocess.Popen, mes: int, ano: int):
    """Monitora processo em background e atualiza progress.json"""
    fase_atual = 0
    inicio_tempo = datetime.now()
    mensagens = []
    
    try:
        # Ler stdout/stderr do processo
        while process.poll() is None:
            # Atualizar progresso baseado no tempo decorrido e fases esperadas
            tempo_decorrido = (datetime.now() - inicio_tempo).total_seconds()
            
            # Estimar fase baseado no tempo (aproximação)
            if tempo_decorrido < 5:
                fase_atual = 0
            elif tempo_decorrido < 15:
                fase_atual = 1
            elif tempo_decorrido < 30:
                fase_atual = 2
            elif tempo_decorrido < 60:
                fase_atual = 3
            elif tempo_decorrido < 120:
                fase_atual = 4
            elif tempo_decorrido < 300:
                fase_atual = 5
            else:
                fase_atual = min(6, fase_atual + 1 if tempo_decorrido % 30 < 5 else fase_atual)
            
            etapa, percent = FASES_CALCULO[min(fase_atual, len(FASES_CALCULO) - 1)]
            
            # Tentar ler linhas do stdout (não bloqueante - apenas verificar se há dados)
            # No Windows, stdout.readline() bloqueia, então apenas verificamos ao final
            try:
                # Verificar se processo ainda está rodando (poll retorna None se estiver)
                if process.poll() is not None:
                    # Processo terminou, ler saída completa
                    break
            except Exception:
                pass
            
            # Atualizar progress.json
            progress_data = {
                "job_id": job_id,
                "percent": percent,
                "etapa": etapa,
                "mensagens": mensagens[-10:],  # Últimas 10 mensagens
                "status": "em_andamento"
            }
            
            try:
                with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
                    json.dump(progress_data, f, ensure_ascii=False)
            except Exception:
                pass
            
            await asyncio.sleep(2)  # Atualizar a cada 2 segundos
        
        # Processo terminou
        return_code = process.returncode
        
        # Não há stdout/stderr para ler pois usamos DEVNULL
        # O código de retorno já indica sucesso ou erro
        
        # Verificar se arquivo de resultado foi gerado
        resultado_path = get_resultado_path()
        
        if return_code == 0 and resultado_path:
            status_final = "concluido"
            etapa_final = "Concluído"
            percent_final = 100
        else:
            status_final = "erro" if return_code != 0 else "concluido"
            etapa_final = f"Processo finalizado (código: {return_code})"
            percent_final = 100
        
        # Atualizar progress.json final
        progress_data = {
            "job_id": job_id,
            "percent": percent_final,
            "etapa": etapa_final,
            "mensagens": mensagens[-20:],
            "status": status_final
        }
        
        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump(progress_data, f, ensure_ascii=False)
        
    except Exception as e:
        # Em caso de erro no monitoramento
        progress_data = {
            "job_id": job_id,
            "percent": 0,
            "etapa": f"Erro no monitoramento: {str(e)}",
            "mensagens": mensagens[-10:],
            "status": "erro"
        }
        try:
            with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
                json.dump(progress_data, f, ensure_ascii=False)
        except Exception:
            pass
    finally:
        # Remover do dicionário de processos ativos
        processos_ativos.pop(job_id, None)

@app.post("/calcular")
async def iniciar_calculo(mes: int = Query(..., ge=1, le=12), ano: int = Query(..., ge=2000, le=2100)):
    """Inicia cálculo de comissões"""
    job_id = str(uuid.uuid4())
    
    # Criar arquivo de progresso inicial
    progress_data = {
        "job_id": job_id,
        "percent": 0,
        "etapa": "Iniciando...",
        "mensagens": [],
        "status": "em_andamento"
    }
    
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress_data, f, ensure_ascii=False)
    
    # Disparar subprocesso
    script_path = Path(ROBO_ROOT_PATH) / "calculo_comissoes.py"
    if not script_path.exists():
        raise HTTPException(status_code=404, detail="Arquivo calculo_comissoes.py não encontrado")
    
    # Iniciar processo em background com parâmetros mes/ano
    # Redirecionar stdout/stderr para DEVNULL para evitar bloqueio por buffers cheios
    # O processo não ficará bloqueado esperando que alguém leia os pipes
    process = subprocess.Popen(
        [sys.executable, str(script_path), "--mes", str(mes), "--ano", str(ano)],
        cwd=str(ROBO_ROOT_PATH),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
    )
    
    # Armazenar processo ativo
    processos_ativos[job_id] = process
    
    # Iniciar monitoramento em background
    asyncio.create_task(monitorar_processo(job_id, process, mes, ano))
    
    return {"job_id": job_id, "message": "Cálculo iniciado"}


@app.get("/progresso/{job_id}")
async def consultar_progresso(job_id: str):
    """Consulta progresso do cálculo"""
    if not os.path.exists(PROGRESS_FILE):
        # Simular progresso mínimo
        return ProgressResponse(
            job_id=job_id,
            percent=0,
            etapa="Aguardando início...",
            mensagens=[],
            status="em_andamento"
        )
    
    try:
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            progress = json.load(f)
        
        # Verificar se job_id corresponde
        if progress.get("job_id") != job_id:
            # Criar novo progresso
            return ProgressResponse(
                job_id=job_id,
                percent=0,
                etapa="Aguardando início...",
                mensagens=[],
                status="em_andamento"
            )
        
        return ProgressResponse(**progress)
    except Exception as e:
        return ProgressResponse(
            job_id=job_id,
            percent=0,
            etapa=f"Erro ao ler progresso: {str(e)}",
            mensagens=[],
            status="erro"
        )


# ==================== ENDPOINTS - RESULTADOS ====================

@app.get("/resultado/abas")
async def listar_abas_resultado():
    """Lista abas do arquivo de resultado mais recente"""
    resultado_path = get_resultado_path()
    if not resultado_path:
        return {"abas": []}
    
    try:
        wb = load_workbook(resultado_path, read_only=True)
        return {"abas": wb.sheetnames, "arquivo": resultado_path.name}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler resultado: {str(e)}")


@app.get("/resultado/aba/{nome_aba}")
async def ler_aba_resultado(
    nome_aba: str,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=1000),
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = Query("asc", regex="^(asc|desc)$"),
    filters: Optional[str] = None,
):
    """Lê uma aba do resultado com paginação"""
    resultado_path = get_resultado_path()
    if not resultado_path:
        raise HTTPException(status_code=404, detail="Nenhum arquivo de resultado encontrado")
    
    df = read_excel_sheet(resultado_path, nome_aba)
    
    # Aplicar filtros
    if filters:
        try:
            filter_dict = json.loads(filters)
            for col, value in filter_dict.items():
                if col in df.columns and value:
                    df = df[df[col].astype(str).str.contains(str(value), case=False, na=False)]
        except Exception:
            pass
    
    # Ordenação
    if sort_by and sort_by in df.columns:
        ascending = sort_order == "asc"
        df = df.sort_values(by=sort_by, ascending=ascending)
    
    # Paginação
    total = len(df)
    start = (page - 1) * size
    end = start + size
    df_page = df.iloc[start:end]
    
    return {
        "data": df_page.to_dict(orient="records"),
        "total": total,
        "page": page,
        "size": size,
        "columns": list(df.columns),
        "arquivo": resultado_path.name
    }


@app.get("/baixar/resultado")
async def baixar_resultado():
    """Download do arquivo de resultado mais recente"""
    resultado_path = get_resultado_path()
    if not resultado_path:
        raise HTTPException(status_code=404, detail="Nenhum arquivo de resultado encontrado")
    
    return FileResponse(
        resultado_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=resultado_path.name
    )


@app.get("/health")
async def health_check():
    """Health check"""
    return {
        "status": "ok",
        "robo_path": ROBO_ROOT_PATH,
        "regras_exists": get_regras_path().exists(),
        "resultado_exists": get_resultado_path() is not None
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

