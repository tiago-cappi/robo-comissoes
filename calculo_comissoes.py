import pandas as pd
import numpy as np
import os
import preparar_dados_mensais
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import calendar
import time
import logging
import unicodedata
import re
import sys

import os, sys

# Flag simples de verbosidade (NÃO muda cálculo)
LOG_VERBOSE = os.getenv("COMISSOES_VERBOSE", "0") == "1"

def _phase(title: str):
    """Imprime separador visual para início de fase/etapa."""
    sep = "#" * 72
    print(f"\n{sep}\n# {title}\n{sep}\n")

def _progress_step5(current: int, total: int, width: int = 30):
    """
    Barra de progresso para Etapa 5 (calculando FC item a item).
    Usa apenas '#######' e percentual, sem mudar a lógica de cálculo.
    """
    if total <= 0:
        total = 1  # evita divisão por zero
    pct = int((current / total) * 100)
    filled = int((pct * width) // 100)
    bar = "#" * filled + "-" * (width - filled)
    sys.stdout.write(f"\r[Etapa 5] calculando FC item a item: [{bar}] {pct}% ({current}/{total})")
    sys.stdout.flush()
    if current >= total:
        sys.stdout.write("\n")

def _info(msg: str):
    """Log leve (sempre mostrado)."""
    print(msg)

def _debug(msg: str):
    """Log verboso (apenas se COMISSOES_VERBOSE=1)."""
    if LOG_VERBOSE:
        print(msg)


# ======== ESTILIZAÇÃO DO EXCEL (pós-processamento, sem mudar dados) ========
def _light_fill(rgb_hex: str) -> PatternFill:
    # rgb_hex sem '#', ex: 'E3F2FD'
    return PatternFill(start_color=rgb_hex, end_color=rgb_hex, fill_type="solid")


# Paleta suave (pastéis, claras)
_PALETTE = [
    "E3F2FD",  # azul claríssimo
    "E8F5E9",  # verde claríssimo
    "FFF8E1",  # amarelo claríssimo
    "F3E5F5",  # lilás claríssimo
    "E0F7FA",  # ciano claríssimo
    "FBE9E7",  # pêssego claríssimo
    "F1F8E9",  # lima claríssimo
    "FFF3E0",  # laranja claríssimo
    "EDE7F6",  # roxo claríssimo
    "F9FBE7",  # chartreuse claríssimo
]


# Grupo → padrões de identificação de coluna (qualquer substring no cabeçalho)
_FC_GROUP_PATTERNS = {
    # faturamento (linha / individual)
    "faturamento_linha": [
        "FATURAMENTO_LINHA",
        "FC_FATURAMENTO_LINHA",
        "PESO_FATURAMENTO_LINHA",
        "META_FATURAMENTO_LINHA",
        "REALIZADO_FATURAMENTO_LINHA",
        "ATINGIMENTO_FATURAMENTO_LINHA",
        "PESO_FAT_LINHA",
        "META_FAT_LINHA",
        "REALIZADO_FAT_LINHA",
        "ATING_FAT_LINHA",
        "ATING_CAP_FAT_LINHA",
        "COMP_FC_FAT_LINHA",
    ],
    "faturamento_individual": [
        "FATURAMENTO_INDIVIDUAL",
        "FC_FATURAMENTO_INDIVIDUAL",
        "PESO_FATURAMENTO_INDIVIDUAL",
        "META_FATURAMENTO_INDIVIDUAL",
        "REALIZADO_FATURAMENTO_INDIVIDUAL",
        "ATINGIMENTO_FATURAMENTO_INDIVIDUAL",
        "PESO_FAT_IND",
        "META_FAT_IND",
        "REALIZADO_FAT_IND",
        "ATING_FAT_IND",
        "ATING_CAP_FAT_IND",
        "COMP_FC_FAT_IND",
    ],

    # conversão (linha / individual)
    "conversao_linha": [
        "CONVERSAO_LINHA",
        "FC_CONVERSAO_LINHA",
        "PESO_CONVERSAO_LINHA",
        "META_CONVERSAO_LINHA",
        "REALIZADO_CONVERSAO_LINHA",
        "ATINGIMENTO_CONVERSAO_LINHA",
        "PESO_CONV_LINHA",
        "META_CONV_LINHA",
        "REALIZADO_CONV_LINHA",
        "ATING_CONV_LINHA",
        "ATING_CAP_CONV_LINHA",
        "COMP_FC_CONV_LINHA",
    ],
    "conversao_individual": [
        "CONVERSAO_INDIVIDUAL",
        "FC_CONVERSAO_INDIVIDUAL",
        "PESO_CONVERSAO_INDIVIDUAL",
        "META_CONVERSAO_INDIVIDUAL",
        "REALIZADO_CONVERSAO_INDIVIDUAL",
        "ATINGIMENTO_CONVERSAO_INDIVIDUAL",
        "PESO_CONV_IND",
        "META_CONV_IND",
        "REALIZADO_CONV_IND",
        "ATING_CONV_IND",
        "ATING_CAP_CONV_IND",
        "COMP_FC_CONV_IND",
    ],

    # rentabilidade / retenção
    "rentabilidade": [
        "RENTABILIDADE",
        "FC_RENTABILIDADE",
        "PESO_RENTABILIDADE",
        "META_RENTABILIDADE",
        "REALIZADO_RENTABILIDADE",
        "ATINGIMENTO_RENTABILIDADE",
        "PESO_RENTAB",
        "META_RENTAB",
        "REALIZADO_RENTAB",
        "ATING_RENTAB",
        "ATING_CAP_RENTAB",
        "COMP_FC_RENTAB",
    ],
    "retencao": [
        "RETENCAO",
        "FC_RETENCAO",
        "PESO_RETENCAO",
        "META_RETENCAO",
        "REALIZADO_RETENCAO",
        "ATINGIMENTO_RETENCAO",
        "PESO_RETENCAO",
        "META_RETENCAO",
        "REALIZADO_RETENCAO",
        "ATING_RETENCAO",
        "ATING_CAP_RETENCAO",
        "COMP_FC_RETENCAO",
    ],

    # metas de fornecedores (se existirem)
    "fornecedor_1": [
        "FORNECEDOR_1",
        "FC_FORNECEDOR_1",
        "PESO_FORNECEDOR_1",
        "META_FORNECEDOR_1",
        "REALIZADO_FORNECEDOR_1",
        "ATINGIMENTO_FORNECEDOR_1",
        "PESO_FORN1",
        "META_FORN1",
        "REALIZADO_FORN1",
        "ATING_FORN1",
        "ATING_CAP_FORN1",
        "COMP_FC_FORN1",
        "MOEDA_FORN1",
    ],
    "fornecedor_2": [
        "FORNECEDOR_2",
        "FC_FORNECEDOR_2",
        "PESO_FORNECEDOR_2",
        "META_FORNECEDOR_2",
        "REALIZADO_FORNECEDOR_2",
        "ATINGIMENTO_FORNECEDOR_2",
        "PESO_FORN2",
        "META_FORN2",
        "REALIZADO_FORN2",
        "ATING_FORN2",
        "ATING_CAP_FORN2",
        "COMP_FC_FORN2",
        "MOEDA_FORN2",
    ],

    # FC total (às vezes reportado em colunas próprias)
    "fc_total": [
        "FC_TOTAL",
        "FC_GERAL",
        "FATOR_CORRECAO_FC",
        "FATOR_CORRECAO",
        "FC",
    ],
}


def _match_group(header: str):
    h = header.upper().strip()
    for group, patterns in _FC_GROUP_PATTERNS.items():
        for p in patterns:
            if p in h:
                return group
    return None


def _apply_group_fills_to_sheet(ws):
    """
    Pinta colunas inteiras do mesmo grupo (componente FC/meta/realizado/peso/atingimento)
    com a mesma cor clara. Não altera valores.
    """
    header_row = 1
    col_group = {}
    max_col = ws.max_column
    max_row = ws.max_row

    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        group = _match_group(str(cell.value) if cell.value is not None else "")
        if group:
            col_group[c] = group

    if not col_group:
        return

    group_keys = sorted(set(col_group.values()))
    color_for_group = {}
    for idx, g in enumerate(group_keys):
        color_for_group[g] = _light_fill(_PALETTE[idx % len(_PALETTE)])

    for c, g in col_group.items():
        fill = color_for_group[g]
        for r in range(1, max_row + 1):
            ws.cell(row=r, column=c).fill = fill


def style_output_workbook(xlsx_path: str):
    """
    Carrega o arquivo Excel gerado e aplica a coloração de grupos
    nas abas COMISSOES_CALCULADAS e RECONCILIACOES/RECONCILIACAO.
    """
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        return

    targets = []
    for name in ("COMISSOES_CALCULADAS", "RECONCILIACOES", "RECONCILIACAO"):
        if name in wb.sheetnames:
            targets.append(name)

    for sheet_name in targets:
        ws = wb[sheet_name]
        _apply_group_fills_to_sheet(ws)

    wb.save(xlsx_path)


def _normalize_text(s):
    if pd.isna(s):
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    return " ".join(s.strip().upper().split())


try:
    import requests
except Exception:
    requests = None

# Tenta importar a biblioteca para PDF. Se não existir, o script funcionará sem gerar o PDF.
try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.units import inch
    REPORTLAB_DISPONIVEL = True
except ImportError:
    REPORTLAB_DISPONIVEL = False

# --- CONFIGURAÇÕES E CONSTANTES ---
# Nomes dos arquivos de entrada (ajuste se necessário)
ARQUIVO_REGRAS_XLSX = "Regras_Comissoes.xlsx"
ARQUIVO_RETENCAO = "Retencao_Clientes.xlsx"
ARQUIVO_RECEBIMENTOS = "Recebimentos_do_Mes.xlsx"
ARQUIVO_STATUS_PAGAMENTOS = "Status_Pagamentos_Processos.xlsx"
ARQUIVO_ESTADO = "Estado_Processos_Recebimento.xlsx"
# FORÇAR DEBUG TEMPORARIAMENTE (será desativado após a execução)
FORCE_DEBUG_TERMINAL = False

# Nome do arquivo de saída
NOME_ARQUIVO_SAIDA = "Comissoes_Calculadas_{}.xlsx".format(datetime.now().strftime('%Y%m%d_%H%M%S'))

class CalculoComissao:
    """
    Classe principal para orquestrar o cálculo de comissões.
    """
    def __init__(self):
        self.data = {}
        self.params = {}
        self.validation_log = []
        self.legacy_token = '__legacy__'
        self.cache_regras = {}
        # Cache para taxas de câmbio: chave (ano, mes_final, tuple(moedas)) -> dict
        self.cache_cambio = {}
        # Coleta de depuração para metas de fornecedores
        self.debug_fornecedores = []
        # Decisões e marcações de cross-selling por Processo
        self.cross_selling_decisions = {}
        # Estruturas auxiliares para conciliação retroativa
        self.reconciliacao_detalhada_list = []
        self.reconciliacao_resumo_list = []
        # Caminho base para localizar arquivos históricos
        self.base_path = os.getcwd()

    def _log_validacao(self, nivel, mensagem, contexto={}):
        """Adiciona uma entrada ao log de validação."""
        self.validation_log.append({
            "Nível": nivel,
            "Mensagem": mensagem,
            "Contexto": str(contexto)
        })

    def _carregar_dados(self):
        """Carrega todos os arquivos de entrada."""
        try:
            regras_data = pd.read_excel(ARQUIVO_REGRAS_XLSX, sheet_name=None)
            self.data.update(regras_data)
            # FATURADOS e CONVERSOES podem não existir para o mês — carregar defensivamente
            # FATURADOS: try the configured file, otherwise try common fallbacks (generic filenames or any matching pattern)
            def _try_read_any(candidates):
                for p in candidates:
                    try:
                        if p and os.path.exists(p):
                            return pd.read_excel(p)
                    except Exception:
                        continue
                # try pattern scan in cwd
                try:
                    for fname in os.listdir('.'):
                        if fname.lower().startswith('faturados') and fname.lower().endswith(('.xls', '.xlsx', '.csv')):
                            try:
                                return pd.read_excel(fname)
                            except Exception:
                                continue
                except Exception:
                    return pd.DataFrame()
                    pass
                return pd.DataFrame()
            self.data['FATURADOS'] = _try_read_any([
            ARQUIVO_FATURADOS,
            'Faturados.xlsx',
            'Faturados.xls',
            'Faturados.csv'
            ])
            self.data['CONVERSOES'] = _try_read_any([ARQUIVO_CONVERSOES, 'Conversões.xlsx', 'Conversoes.xlsx', 'Conversões.csv', 'Conversoes.csv'])
            self.data['RENTABILIDADE_REALIZADA'] = pd.read_excel(ARQUIVO_RENTABILIDADE)
            # Novo arquivo com dados de retenção de clientes por linha
            try:
                self.data['RETENCAO_CLIENTES'] = pd.read_excel(ARQUIVO_RETENCAO)
            except FileNotFoundError:
                # Se o arquivo não existir, deixamos a chave com DataFrame vazio
                self.data['RETENCAO_CLIENTES'] = pd.DataFrame(columns=['linha', 'clientes_mes_anterior', 'clientes_mes_atual'])

            # Carrega o faturamento YTD (contendo coluna 'Fabricante' e datas completas)
            try:
                self.data['FATURADOS_YTD'] = pd.read_excel(ARQUIVO_FATURADOS_YTD, parse_dates=['Dt Emissão'])
            except FileNotFoundError:
                self.data['FATURADOS_YTD'] = pd.DataFrame(columns=['Dt Emissão', 'Fabricante', 'Valor Realizado'])

            # Nova aba com metas por fornecedor
            if 'METAS_FORNECEDORES' in self.data:
                self.data['METAS_FORNECEDORES'] = self.data['METAS_FORNECEDORES']
            else:
                # Se não existir a aba, criamos estrutura vazia para não quebrar o fluxo
                self.data['METAS_FORNECEDORES'] = pd.DataFrame(columns=['linha', 'fornecedor', 'meta_anual', 'moeda'])

            # Nova aba CROSS_SELLING (opcional): colaborador (Consultor Externo) e taxa_cross_selling_pct
            try:
                self.data['CROSS_SELLING'] = pd.read_excel(ARQUIVO_REGRAS_XLSX, sheet_name='CROSS_SELLING')
            except Exception:
                # se não existir, manter DataFrame vazio com colunas esperadas
                self.data['CROSS_SELLING'] = pd.DataFrame(columns=['colaborador', 'taxa_cross_selling_pct'])

            # Normalização segura de colunas/strings (evita .str em valores não-string)
            for df_name, df_any in list(self.data.items()):
                try:
                    if not isinstance(df_any, pd.DataFrame):
                        continue
                    try:
                        df_any.columns = df_any.columns.astype(str).str.strip()
                    except Exception:
                        pass
                    for col in df_any.columns:
                        s = df_any[col]
                        if pd.api.types.is_object_dtype(s) or pd.api.types.is_string_dtype(s):
                            try:
                                df_any[col] = s.apply(lambda v: v.strip() if isinstance(v, str) else v)
                            except Exception:
                                pass
                    self.data[df_name] = df_any
                except Exception as _e_norm:
                    self._log_validacao('AVISO', f'Falha ao normalizar strings para {df_name}: {_e_norm}', {})

            # Carregar recebimentos do mês (opcional)
            try:
                self.data['RECEBIMENTOS'] = pd.read_excel(ARQUIVO_RECEBIMENTOS)
            except Exception:
                self.data['RECEBIMENTOS'] = pd.DataFrame(columns=['PROCESSO', 'DATA_RECEBIMENTO', 'VALOR_RECEBIDO', 'ID_CLIENTE'])

            # Carregar arquivo de análise comercial completo (histórico de processos)
            try:
                # suportar .xlsx, .xls ou .csv (usuário informou que tem .csv)
                analise_candidates = [
                    'Analise_Comercial_Completa.xlsx',
                    'Analise_Comercial_Completa.xls',
                    'Analise_Comercial_Completa.csv'
                ]
                analise_path = None
                for p in analise_candidates:
                    if os.path.exists(p):
                        analise_path = p
                        break

                if analise_path is None:
                    # manter DataFrame vazio — mais adiante a geração de COMISSOES_RECEBIMENTO
                    # exige explicitamente este arquivo; registramos para validação
                    self.data['ANALISE_COMERCIAL_COMPLETA'] = pd.DataFrame()
                    self._log_validacao('AVISO', 'ANALISE_COMERCIAL_COMPLETA ausente (procurados .xlsx/.xls/.csv).', {'candidates': analise_candidates})
                if analise_path is not None:
                    # ler CSV ou Excel conforme a extensão
                    try:
                        if analise_path.lower().endswith('.csv'):
                            # tentar detectar delimitador automaticamente
                            df_anal = pd.read_csv(analise_path, sep=None, engine='python', dtype=str)
                        else:
                            # Excel: inferir se existe coluna 'Dt Emissão' para parse_dates
                            try:
                                hdrs = pd.read_excel(analise_path, nrows=0).columns.tolist()
                            except Exception:
                                hdrs = []
                            parse_dates = ['Dt Emissão'] if 'Dt Emissão' in hdrs else False
                            df_anal = pd.read_excel(analise_path, parse_dates=parse_dates, dtype=str)
                    except Exception as e_read:
                        self._log_validacao('AVISO', f'Falha ao ler {analise_path}: {e_read}', {'path': analise_path})
                        df_anal = pd.DataFrame()

                    # Normalizar colunas e strings (trim)
                    if not df_anal.empty:
                        df_anal.columns = df_anal.columns.str.strip()
                        for c in df_anal.select_dtypes(include=['object']):
                            df_anal[c] = df_anal[c].astype(str).str.strip()

                    self.data['ANALISE_COMERCIAL_COMPLETA'] = df_anal
                    if getattr(self, '_logger', None):
                        self._logger.info(f"ANALISE_COMERCIAL_COMPLETA carregada de: {analise_path} (linhas={len(df_anal)})")
            except Exception as e:
                self.data['ANALISE_COMERCIAL_COMPLETA'] = pd.DataFrame()
                self._log_validacao('AVISO', f'Erro ao inicializar ANALISE_COMERCIAL_COMPLETA: {e}', {})

            # Carregar status de pagamentos (opcional)
            try:
                self.data['STATUS_PAGAMENTOS'] = pd.read_excel(ARQUIVO_STATUS_PAGAMENTOS)
            except Exception:
                self.data['STATUS_PAGAMENTOS'] = pd.DataFrame(columns=['PROCESSO', 'VALOR_ORIGINAL', 'STATUS_PAGAMENTO'])

            # Normalizar colunas da aba METAS_FORNECEDORES: alguns arquivos usam 'fabricante' em vez de 'fornecedor'
            if 'METAS_FORNECEDORES' in self.data:
                df_met = self.data['METAS_FORNECEDORES']
                if 'fabricante' in df_met.columns and 'fornecedor' not in df_met.columns:
                    df_met = df_met.rename(columns={'fabricante': 'fornecedor'})
                    self.data['METAS_FORNECEDORES'] = df_met

            params_df = self.data['PARAMS']
            self.params = pd.Series(params_df.valor.values, index=params_df.chave).to_dict()
            # parâmetro para escolha default em execuções não interativas
            self.params['cross_selling_default_option'] = str(self.params.get('cross_selling_default_option', 'A')).upper()
            param_base_path = self.params.get('base_path')
            if param_base_path:
                self.base_path = str(param_base_path)
            self.legacy_token = self.params.get('legacy_scope_token', '__legacy__')
            # Configurar logger para depuração no terminal se solicitado
            debug_terminal = str(self.params.get('debug_terminal_fornecedores', False)).lower() in ('1', 'true', 'yes')
            # Forçar DEBUG em execução temporária se solicitado pelo desenvolvedor
            debug_terminal = debug_terminal or FORCE_DEBUG_TERMINAL
            self._logger = logging.getLogger('calculo_comissoes')
            handler = logging.StreamHandler(sys.stdout)
            formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s')
            handler.setFormatter(formatter)
            if not self._logger.handlers:
                self._logger.addHandler(handler)
            self._logger.propagate = False
            self._logger.setLevel(logging.DEBUG if debug_terminal else logging.INFO)
        except Exception as e:
            # Erro crítico ao carregar dados iniciais; registrar mas continuar (dados poderão estar incompletos)
            self._log_validacao('ERRO', f'Falha ao carregar dados iniciais: {e}', {})

        # Determinar quem recebe por recebimento (colaboradores cujo tipo de comissão é 'Recebimento' ou cargo marcado)
        try:
            recebe_set = set()
            # Prefer explicit coluna TIPO_COMISSAO na aba CARGOS ou COLABORADORES
            if 'CARGOS' in self.data and 'TIPO_COMISSAO' in self.data['CARGOS'].columns:
                cargos_rc = self.data['CARGOS'][self.data['CARGOS']['TIPO_COMISSAO'].astype(str).str.strip().str.lower() == 'recebimento']['nome_cargo'].tolist()
                if getattr(self, '_logger', None):
                    self._logger.info(f"CARGOS marcados como recebimento: {cargos_rc}")
                if cargos_rc and 'COLABORADORES' in self.data:
                    dfcol = self.data['COLABORADORES']
                    colabs_receb = dfcol[dfcol['cargo'].isin(cargos_rc)]['nome_colaborador'].dropna().astype(str).str.strip().tolist()
                    if getattr(self, '_logger', None):
                        self._logger.info(f"Colaboradores detectados para recebimento (via cargo): {colabs_receb}")
                    recebe_set.update(colabs_receb)

            # Se existir coluna 'TIPO_COMISSAO' em COLABORADORES, use-a direto
            if 'COLABORADORES' in self.data and 'TIPO_COMISSAO' in self.data['COLABORADORES'].columns:
                colabs_receb2 = self.data['COLABORADORES'][self.data['COLABORADORES']['TIPO_COMISSAO'].astype(str).str.strip().str.lower() == 'recebimento']['nome_colaborador'].dropna().astype(str).str.strip().tolist()
                if getattr(self, '_logger', None):
                    self._logger.info(f"Colaboradores detectados para recebimento (via TIPO_COMISSAO): {colabs_receb2}")
                recebe_set.update(colabs_receb2)

            # fallback heurístico: cargos com nome contendo 'Receb' ou 'Recebimento'
            if not recebe_set and 'CARGOS' in self.data and 'nome_cargo' in self.data['CARGOS'].columns:
                heur = self.data['CARGOS'][self.data['CARGOS']['nome_cargo'].astype(str).str.contains('receb', case=False, na=False)]['nome_cargo'].tolist()
                if getattr(self, '_logger', None):
                    self._logger.info(f"CARGOS detectados por heurística (nome contém 'receb'): {heur}")
                if heur and 'COLABORADORES' in self.data:
                    colabs_heur = self.data['COLABORADORES'][self.data['COLABORADORES']['cargo'].isin(heur)]['nome_colaborador'].dropna().astype(str).str.strip().tolist()
                    if getattr(self, '_logger', None):
                        self._logger.info(f"Colaboradores detectados por heurística: {colabs_heur}")
                    recebe_set.update(colabs_heur)

            self.recebe_por_recebimento = set(recebe_set)
            if getattr(self, '_logger', None):
                self._logger.info(f"Set final de colaboradores que recebem por recebimento: {self.recebe_por_recebimento}")
        except FileNotFoundError as e:
            raise Exception(f"Erro: Arquivo não encontrado: {e.filename}.")
        except Exception:
            self.recebe_por_recebimento = set()

    def _validar_dados(self):
        """Executa validações básicas nos dados carregados."""
        # Atualizado para incluir 'retencao_clientes' e metas de fornecedor quando presentes
        peso_cols = ['faturamento_linha', 'rentabilidade', 'conversao_linha', 'faturamento_individual', 'conversao_individual']
        if 'retencao_clientes' in self.data['PESOS_METAS'].columns:
            peso_cols = peso_cols + ['retencao_clientes']
        # Incluir pesos para metas de fornecedores se existirem
        if 'meta_fornecedor_1' in self.data['PESOS_METAS'].columns:
            peso_cols = peso_cols + ['meta_fornecedor_1']
        if 'meta_fornecedor_2' in self.data['PESOS_METAS'].columns:
            peso_cols = peso_cols + ['meta_fornecedor_2']
        pesos_por_cargo = self.data['PESOS_METAS'].groupby('cargo')[peso_cols].sum().sum(axis=1)
        for cargo, soma in pesos_por_cargo.items():
            if not np.isclose(soma, 100) and not np.isclose(soma, 0):
                self._log_validacao('AVISO', f"A soma dos pesos para o cargo '{cargo}' não é 100% (soma: {soma}).", {"cargo": cargo})
        
        colabs_regras = set(self.data['COLABORADORES']['nome_colaborador'])
        colabs_atribuicoes = set(self.data['ATRIBUICOES']['colaborador'])
        nao_encontrados = colabs_atribuicoes - colabs_regras
        for colab in nao_encontrados:
            self._log_validacao('ERRO', f"Colaborador '{colab}' das atribuições não encontrado na lista de colaboradores.", {"colaborador": colab})

        # Registrar quais linhas possuem metas de fornecedores (útil para depuração)
        metas_fornecedores_df = self.data.get('METAS_FORNECEDORES', pd.DataFrame())
        if not metas_fornecedores_df.empty:
            linhas_com_metas = sorted(metas_fornecedores_df['linha'].dropna().unique().tolist())
            self._log_validacao('INFO', f"Linhas com METAS_FORNECEDORES encontradas: {linhas_com_metas}", {'linhas': linhas_com_metas})

    def _preprocessar_dados(self):
        """Prepara os dados para o cálculo, aplicando aliases e conversões."""
        alias_map = self.data['ALIASES'][self.data['ALIASES']['entidade'] == 'colaborador'].set_index('alias')['padrao'].to_dict()
        
        # Aplica alias para todas as colunas de consultores nos arquivos de ERP
        for df_name in ['FATURADOS', 'CONVERSOES']:
            df = self.data.get(df_name, pd.DataFrame())
            # Garantir colunas mínimas para não quebrar o fluxo quando o arquivo estiver ausente ou vazio
            if 'Consultor Interno' not in df.columns:
                df['Consultor Interno'] = ''
            if 'Representante-pedido' not in df.columns:
                df['Representante-pedido'] = ''
            # Aplicar mapeamento de aliases com fallback seguro
            try:
                df['Consultor Interno'] = df['Consultor Interno'].astype(str).replace(alias_map).str.strip()
            except Exception:
                df['Consultor Interno'] = df['Consultor Interno'].astype(str).str.strip()
            try:
                df['Representante-pedido'] = df['Representante-pedido'].astype(str).replace(alias_map).str.strip()
            except Exception:
                df['Representante-pedido'] = df['Representante-pedido'].astype(str).str.strip()
            self.data[df_name] = df

        self.data['COLABORADORES'] = self.data['COLABORADORES'].merge(
            self.data['CARGOS'], left_on='cargo', right_on='nome_cargo', how='left'
        )

    def _calcular_realizado(self):
        """Calcula os valores realizados para faturamento, conversão e rentabilidade."""
        self.realizado = {}
        # FATURADOS: garantir colunas esperadas e agregar com segurança
        df_fat = self.data.get('FATURADOS', pd.DataFrame()).copy()
        if 'Valor Realizado' not in df_fat.columns:
            df_fat['Valor Realizado'] = 0.0
        if 'Negócio' not in df_fat.columns:
            df_fat['Negócio'] = ''
        if 'Consultor Interno' not in df_fat.columns:
            df_fat['Consultor Interno'] = ''
        self.realizado['faturamento_linha'] = df_fat.groupby('Negócio')['Valor Realizado'].sum() if not df_fat.empty else pd.Series(dtype=float)
        self.realizado['faturamento_individual'] = df_fat.groupby('Consultor Interno')['Valor Realizado'].sum() if not df_fat.empty else pd.Series(dtype=float)

        # CONVERSOES: garantir colunas esperadas e agregar com segurança
        df_conv = self.data.get('CONVERSOES', pd.DataFrame()).copy()
        if 'Valor Orçado' not in df_conv.columns and 'Valor Orcado' in df_conv.columns:
            # lidar com variação possível de acentuação
            df_conv['Valor Orçado'] = df_conv['Valor Orcado']
        if 'Valor Orçado' not in df_conv.columns:
            df_conv['Valor Orçado'] = 0.0
        if 'Negócio' not in df_conv.columns:
            df_conv['Negócio'] = ''
        if 'Consultor Interno' not in df_conv.columns:
            df_conv['Consultor Interno'] = ''
        self.realizado['conversao_linha'] = df_conv.groupby('Negócio')['Valor Orçado'].sum() if not df_conv.empty else pd.Series(dtype=float)
        self.realizado['conversao_individual'] = df_conv.groupby('Consultor Interno')['Valor Orçado'].sum() if not df_conv.empty else pd.Series(dtype=float)
        rent_realizada = self.data['RENTABILIDADE_REALIZADA'].rename(columns={'Negócio': 'linha'})
        self.realizado['rentabilidade'] = rent_realizada.set_index(['linha', 'Grupo', 'Subgrupo', 'Tipo de Mercadoria'])['rentabilidade_realizada_pct']

    def _get_meta(self, tipo_meta, chave):
        """Busca o valor da meta correspondente."""
        try:
            if tipo_meta in ['faturamento_linha', 'conversao_linha']:
                df = self.data['METAS_APLICACAO']
                tipo_meta_busca = tipo_meta.replace('_linha', '')
                linha, tipo_mercadoria = chave
                valor = df[(df['linha'] == linha) & (df['tipo_mercadoria'] == tipo_mercadoria) & (df['tipo_meta'] == tipo_meta_busca)]['valor_meta'].iloc[0]
                return valor
            elif tipo_meta in ['faturamento_individual', 'conversao_individual']:
                df = self.data['METAS_INDIVIDUAIS']
                tipo_meta_busca = tipo_meta.replace('_individual', '')
                valor = df[(df['colaborador'] == chave) & (df['tipo_meta'] == tipo_meta_busca)]['valor_meta'].iloc[0]
                return valor
            elif tipo_meta == 'rentabilidade':
                df = self.data['META_RENTABILIDADE']
                linha, grupo, subgrupo, tipo_mercadoria = chave
                valor = df[(df['linha'] == linha) & (df['grupo'] == grupo) & (df['subgrupo'] == subgrupo) & (df['tipo_mercadoria'] == tipo_mercadoria)]['meta_rentabilidade_alvo_pct'].iloc[0]
                return valor
        except (IndexError, KeyError):
            self._log_validacao('AVISO', f"Meta não encontrada para tipo '{tipo_meta}' e chave '{chave}'.", {'tipo_meta': tipo_meta, 'chave': chave})
            return None
        return None

    def _calcular_fc_para_item(self, nome_colab, cargo_colab, item_faturado):
        """Calcula um FC único para um colaborador e um item faturado específico."""
        pesos = self.data['PESOS_METAS'][self.data['PESOS_METAS']['cargo'] == cargo_colab]
        if pesos.empty:
            return 0, {}
        pesos = pesos.iloc[0]
        fc_total_item = 0

        # Estrutura para coletar detalhes por componente do FC
        detalhes_fc = {}

        item_context = {
            'linha': item_faturado['Negócio'], 'grupo': item_faturado['Grupo'],
            'subgrupo': item_faturado['Subgrupo'], 'tipo_mercadoria': item_faturado['Tipo de Mercadoria']
        }

        metas_config = {
            'faturamento_linha': ('faturamento_linha', (item_context['linha'], item_context['tipo_mercadoria'])),
            'conversao_linha': ('conversao_linha', (item_context['linha'], item_context['tipo_mercadoria'])),
            'faturamento_individual': ('faturamento_individual', nome_colab),
            'conversao_individual': ('conversao_individual', nome_colab),
            'rentabilidade': ('rentabilidade', tuple(item_context.values()))
        }

        for tipo_meta, (realizado_key, meta_chave) in metas_config.items():
            peso = pesos.get(tipo_meta, 0) / 100.0
            if peso == 0:
                continue
            
            if tipo_meta.endswith('_linha'):
                realizado = self.realizado[realizado_key].get(item_context['linha'], 0)
            elif tipo_meta.endswith('_individual'):
                realizado = self.realizado[realizado_key].get(nome_colab, 0)
            else: # rentabilidade
                realizado = self.realizado[realizado_key].get(meta_chave, 0)
                # garantir que realizado de rentabilidade esteja em decimal (ex: 0.12)
                try:
                    if realizado is not None:
                        rv = float(realizado)
                        if rv > 1:
                            rv = rv / 100.0
                        realizado = rv
                except Exception:
                    pass

            meta = self._get_meta(tipo_meta, meta_chave)
            atingimento = (realizado / meta) if meta and meta > 0 else 0
            
            cap_atingimento = float(self.params.get('cap_atingimento_max', 1.0))
            atingimento_cap = min(atingimento, cap_atingimento)
            componente_fc = atingimento_cap * peso
            fc_total_item += componente_fc

            # armazenar detalhe deste componente
            detalhes_fc[tipo_meta] = {
                'peso': peso,
                'realizado': realizado,
                'meta': meta,
                'atingimento': atingimento,
                'atingimento_cap': atingimento_cap,
                'componente_fc': componente_fc
            }
            
        cap_fc = float(self.params.get('cap_fc_max', 1.0))
        # --- Novo componente: Retenção de Clientes (aplica-se apenas a Gerente Linha) ---
        try:
            if cargo_colab == 'Gerente Linha':
                # Identificar a(s) linha(s) que o gerente é responsável a partir de ATRIBUICOES
                df_atr = self.data.get('ATRIBUICOES', pd.DataFrame())
                linhas_do_gerente = df_atr[df_atr['colaborador'] == nome_colab]['linha'].dropna().unique()
                # Se houver pelo menos uma linha atribuída, usamos a primeira para retenção
                if len(linhas_do_gerente) > 0 and 'RETENCAO_CLIENTES' in self.data:
                    linha_gerente = linhas_do_gerente[0]
                    df_ret = self.data.get('RETENCAO_CLIENTES', pd.DataFrame())
                    # Filtra pela linha
                    ret_row = df_ret[df_ret['linha'] == linha_gerente]
                    if not ret_row.empty:
                        clientes_ant = ret_row.iloc[0].get('clientes_mes_anterior', None)
                        clientes_atual = ret_row.iloc[0].get('clientes_mes_atual', None)
                        # Tratamento: divisão por zero
                        try:
                            if clientes_ant is None or pd.isna(clientes_ant) or float(clientes_ant) == 0:
                                taxa_retencao = 0.0
                            else:
                                taxa_retencao = float(clientes_atual) / float(clientes_ant)
                        except Exception:
                            taxa_retencao = 0.0

                        # Peso da meta para retenção (em % na tabela PESOS_METAS)
                        peso_ret = 0.0
                        pesos_df = self.data.get('PESOS_METAS', pd.DataFrame())
                        if not pesos_df.empty and 'retencao_clientes' in pesos_df.columns:
                            # procura linha pelo cargo
                            row_peso = pesos_df[pesos_df['cargo'] == cargo_colab]
                            if not row_peso.empty:
                                peso_ret = float(row_peso.iloc[0].get('retencao_clientes', 0)) / 100.0

                        cap_atingimento = float(self.params.get('cap_atingimento_max', 1.0))
                        atingimento_cap = min(taxa_retencao, cap_atingimento)
                        componente_fc_ret = atingimento_cap * peso_ret
                        fc_total_item += componente_fc_ret

                        detalhes_fc['retencao_clientes'] = {
                            'peso': peso_ret,
                            'realizado': clientes_atual,
                            'meta': clientes_ant,
                            'atingimento': taxa_retencao,
                            'atingimento_cap': atingimento_cap,
                            'componente_fc': componente_fc_ret
                        }
        except Exception:
            # Em caso de qualquer erro nessa extensão, não interrompemos o cálculo principal
            pass

        # --- Novos componentes: metas por fornecedor (meta_fornecedor_1, meta_fornecedor_2) ---
        try:
            # Buscar metas de fornecedores para a linha do item
            linha_do_item = item_faturado.get('Negócio')
            metas_fornecedores_df = self.data.get('METAS_FORNECEDORES', pd.DataFrame())
            if not metas_fornecedores_df.empty and linha_do_item is not None:
                metas_da_linha = metas_fornecedores_df[metas_fornecedores_df['linha'] == linha_do_item]
                # Esperamos no máximo dois fornecedores listados; iteramos e mapeamos para meta_fornecedor_1/2
                fornecedores = metas_da_linha.to_dict('records')

                # Se não houver metas de fornecedores para essa linha, adicionar debug entry para rastreio
                if len(fornecedores) == 0:
                    # Por padrão não poluir a aba de debug com linhas sem metas.
                    # Se for necessário ver esses casos, ative o parâmetro 'debug_show_missing_fornecedores' em PARAMS
                    show_missing = str(self.params.get('debug_show_missing_fornecedores', False)).lower() in ('1', 'true', 'yes')
                    if show_missing:
                        self.debug_fornecedores.append({
                            'colaborador': nome_colab,
                            'cargo': cargo_colab,
                            'cod_produto': item_faturado.get('Código Produto', None),
                            'linha_item': linha_do_item,
                            'observacao': 'nenhuma_meta_fornecedor_na_linha',
                            'detalhe': 'METAS_FORNECEDORES vazia para esta linha'
                        })

                # Determinar mês de apuração a partir de 'Dt Emissão' do item_faturado, se disponível
                mes_apuracao = None
                dt_emissao = item_faturado.get('Dt Emissão') if 'Dt Emissão' in item_faturado.index else None
                if pd.notna(dt_emissao):
                    try:
                        mes_apuracao = pd.to_datetime(dt_emissao).month
                    except Exception:
                        mes_apuracao = None

                # Se não encontrarmos mês, usamos mês atual
                if mes_apuracao is None:
                    mes_apuracao = datetime.now().month

                # Preparar lista de moedas necessárias para busca de câmbio
                moedas_necessarias = set()
                for f in fornecedores[:2]:
                    moeda = f.get('moeda')
                    if moeda:
                        moedas_necessarias.add(moeda)

                # Buscar taxas de câmbio para o ano corrente
                ano_corrente = pd.to_datetime(item_faturado.get('Dt Emissão', datetime.now())).year
                taxas = self._get_taxas_de_cambio(ano_corrente, mes_apuracao, list(moedas_necessarias)) if moedas_necessarias else {}

                # Para cada fornecedor (até 2), calculamos o componente
                for idx, fornecedor in enumerate(fornecedores[:2], start=1):
                    fornecedor_nome = fornecedor.get('fornecedor')
                    meta_anual = fornecedor.get('meta_anual')
                    moeda = fornecedor.get('moeda')
                    # Inicializar logger (pode não existir)
                    logger = getattr(self, '_logger', None)
                    # Logs detalhados para depuração no terminal
                    try:
                        if logger and logger.isEnabledFor(logging.DEBUG):
                            logger.debug(f"Iniciando cálculo fornecedor#{idx} para colaborador={nome_colab} cargo={cargo_colab} linha={linha_do_item} fornecedor={fornecedor_nome} moeda={moeda} meta_anual={meta_anual}")
                    except Exception:
                        pass
                    if meta_anual is None or fornecedor_nome is None:
                        continue

                    # meta YTD proporcional
                    try:
                        meta_ytd = (float(meta_anual) / 12.0) * float(mes_apuracao)
                    except Exception:
                        meta_ytd = 0.0

                    # Calcular faturamento realizado YTD para este fabricante/fornecedor
                    faturados_ytd = self.data.get('FATURADOS_YTD', pd.DataFrame())
                    # Filtra por fabricante/fornecedor; assumimos coluna 'Fabricante' corresponde ao fornecedor
                    if faturados_ytd.empty:
                        faturamento_realizado_ytd = 0.0
                    else:
                        filt = (faturados_ytd['Fabricante'] == fornecedor_nome)
                        vendas_fornecedor = faturados_ytd[filt].copy()
                        if 'Dt Emissão' in vendas_fornecedor.columns:
                            vendas_fornecedor['mes'] = vendas_fornecedor['Dt Emissão'].dt.month
                        else:
                            # Tentar inferir mês a partir de outra coluna ou assumir todo em mes_apuracao
                            vendas_fornecedor['mes'] = mes_apuracao

                        faturamento_realizado_ytd = 0.0
                        for mes in range(1, mes_apuracao + 1):
                            vendas_do_mes = vendas_fornecedor[vendas_fornecedor['mes'] == mes]
                            soma_brl = vendas_do_mes['Valor Realizado'].sum() if not vendas_do_mes.empty else 0.0
                            taxa_mes = None
                            if moeda and taxas and moeda in taxas and mes in taxas[moeda]:
                                taxa_mes = taxas[moeda].get(mes)
                            # Se taxa_mes for None ou zero, evitamos conversão e consideramos 0 convertido
                            if taxa_mes and taxa_mes != 0:
                                # taxa_mes é a taxa no formato (moeda por 1 BRL),
                                # ou seja: 1 BRL = taxa_mes * MOEDA. Para converter
                                # soma_brl (em BRL) para a moeda alvo, multiplicamos.
                                faturamento_convertido = float(soma_brl) * float(taxa_mes)
                            else:
                                faturamento_convertido = 0.0
                            # Log mensal detalhado
                            try:
                                if logger and logger.isEnabledFor(logging.DEBUG):
                                    logger.debug(f"Fornecedor#{idx} mes={mes}: soma_brl={soma_brl:.2f} taxa_mes={taxa_mes} faturamento_convertido={faturamento_convertido:.4f}")
                            except Exception:
                                pass
                            faturamento_realizado_ytd += faturamento_convertido

                    # Cálculo do atingimento e componente
                    try:
                        atingimento = (faturamento_realizado_ytd / meta_ytd) if meta_ytd and meta_ytd > 0 else 0.0
                    except Exception:
                        atingimento = 0.0


                    cap_atingimento = float(self.params.get('cap_atingimento_max', 1.0))
                    atingimento_cap = min(atingimento, cap_atingimento)

                    # Peso referente a meta_fornecedor_1 ou meta_fornecedor_2 conforme idx
                    peso_col_name = f'meta_fornecedor_{idx}'
                    peso_fornecedor = 0.0
                    pesos_df = self.data.get('PESOS_METAS', pd.DataFrame())
                    if not pesos_df.empty and peso_col_name in pesos_df.columns:
                        row_peso = pesos_df[pesos_df['cargo'] == cargo_colab]
                        if not row_peso.empty:
                            peso_fornecedor = float(row_peso.iloc[0].get(peso_col_name, 0)) / 100.0


                    componente_fc_forn = atingimento_cap * peso_fornecedor

                    # Log resumo do fornecedor
                    try:
                        if logger and logger.isEnabledFor(logging.DEBUG):
                            logger.debug(f"Resumo fornecedor#{idx} colaborador={nome_colab} fornecedor={fornecedor_nome} meta_ytd={meta_ytd:.2f} faturamento_realizado_ytd={faturamento_realizado_ytd:.4f} atingimento={atingimento:.4f} atingimento_cap={atingimento_cap:.4f} peso={peso_fornecedor:.4f} componente_fc={componente_fc_forn:.6f}")
                    except Exception:
                        pass
                    # Coleta de depuração para este cálculo de fornecedor
                    # Sanity-check: recompute converted faturamento from the ON-DISK FATURADOS_YTD file and taxas
                    try:
                        safe_total = 0.0
                        # try to reload the source FATURADOS_YTD from disk to avoid mutated in-memory frames
                        try:
                            faturados_ytd_disk = pd.read_excel(ARQUIVO_FATURADOS_YTD)
                        except Exception:
                            # fallback: use whatever is in memory
                            faturados_ytd_disk = self.data.get('FATURADOS_YTD', pd.DataFrame())

                        # find fabricante and valor columns
                        if not faturados_ytd_disk.empty:
                            fab_col = next((c for c in faturados_ytd_disk.columns if 'fabricante' in str(c).lower() or 'fornecedor' in str(c).lower()), None)
                            val_col = next((c for c in faturados_ytd_disk.columns if 'valor' in str(c).lower()), None)
                            dt_col_local = next((c for c in faturados_ytd_disk.columns if 'dt' in str(c).lower() or 'data' in str(c).lower()), None)
                            if fab_col and val_col:
                                # filter case-insensitive
                                filt_disk = faturados_ytd_disk[faturados_ytd_disk[fab_col].astype(str).str.upper().str.contains(str(fornecedor_nome).upper(), na=False)].copy()
                                if dt_col_local and dt_col_local in filt_disk.columns:
                                    filt_disk['mes'] = pd.to_datetime(filt_disk[dt_col_local], errors='coerce').dt.month
                                else:
                                    filt_disk['mes'] = mes_apuracao
                                for m in range(1, int(mes_apuracao) + 1):
                                    soma_brl = filt_disk[filt_disk['mes'] == m][val_col].sum() if not filt_disk.empty else 0.0
                                    taxa_m = None
                                    if moeda and taxas and moeda in taxas and m in taxas[moeda]:
                                        taxa_m = taxas[moeda].get(m)
                                    if taxa_m and taxa_m != 0:
                                        safe_total += float(soma_brl) * float(taxa_m)
                    except Exception:
                        safe_total = faturamento_realizado_ytd

                    # If the computed total differs wildly from the runtime value (or runtime value is absurd), prefer the recomputed safe_total
                    try:
                        if (faturamento_realizado_ytd is None) or (abs(faturamento_realizado_ytd) > 1e6) or (abs(faturamento_realizado_ytd - safe_total) > max(1e3, abs(0.1 * safe_total))):
                            # Log a validation warning and replace the value to avoid propagation of absurd numbers
                            self._log_validacao('AVISO', f'Valor de faturamento_realizado_ytd anômalo detectado para fornecedor {fornecedor_nome} (orig={faturamento_realizado_ytd}, recomputed={safe_total}). Substituindo pelo valor recomputado.', {'fornecedor': fornecedor_nome, 'orig': faturamento_realizado_ytd, 'recomputed': safe_total})
                            faturamento_realizado_ytd = safe_total
                            # recompute atingimento and component after replacement
                            try:
                                atingimento = (faturamento_realizado_ytd / meta_ytd) if meta_ytd and meta_ytd > 0 else 0.0
                            except Exception:
                                atingimento = 0.0
                            atingimento_cap = min(atingimento, cap_atingimento)
                            componente_fc_forn = atingimento_cap * peso_fornecedor
                    except Exception:
                        pass

                    debug_entry = {
                        'colaborador': nome_colab,
                        'cargo': cargo_colab,
                        'cod_produto': item_faturado.get('Código Produto', None),
                        'linha_item': linha_do_item,
                        'fornecedor_index': idx,
                        'fornecedor': fornecedor_nome,
                        'moeda': moeda,
                        'meta_anual': meta_anual,
                        'meta_ytd': meta_ytd,
                        'faturamento_realizado_ytd': faturamento_realizado_ytd,
                        'mes_apuracao': mes_apuracao,
                        'peso_col_name': peso_col_name,
                        'peso_fornecedor': peso_fornecedor,
                        'atingimento': atingimento,
                        'atingimento_cap': atingimento_cap,
                        'componente_fc': componente_fc_forn
                    }
                    # Observações sobre taxas usadas (se houver)
                    taxas_obs = {}
                    if moeda and 'taxas' in locals() and isinstance(taxas, dict) and moeda in taxas:
                        taxas_obs = taxas.get(moeda, {})
                    debug_entry['taxas_usadas'] = str(taxas_obs)
                    # Indica se houve meses sem taxa (None)
                    taxas_meses_none = [m for m, v in (taxas_obs.items() if isinstance(taxas_obs, dict) else []) if v is None]
                    debug_entry['taxas_meses_none'] = str(taxas_meses_none)
                    debug_entry['taxas_completas'] = (len(taxas_meses_none) == 0) if isinstance(taxas_obs, dict) and len(taxas_obs) > 0 else False
                    self.debug_fornecedores.append(debug_entry)

                    # armazenar detalhes do fornecedor (meta_fornecedor_1/2)
                    detalhes_fc[peso_col_name] = {
                        'peso': peso_fornecedor,
                        'realizado': faturamento_realizado_ytd,
                        'meta': meta_ytd,
                        'atingimento': atingimento,
                        'atingimento_cap': atingimento_cap,
                        'componente_fc': componente_fc_forn,
                        'moeda': moeda
                    }
        except Exception as e:
            # Não interromper fluxo principal em caso de erro nos componentes de fornecedor
            self._log_validacao('AVISO', f'Erro ao calcular metas de fornecedores: {e}', {'item': item_faturado.get('Código Produto', None)})

        return min(fc_total_item, cap_fc), detalhes_fc

    def _get_taxas_de_cambio(self, ano, mes_final, moedas):
        """Retorna dicionário {moeda: {mes: taxa}} com média mensal de cada moeda do mês 1..mes_final.

        Estratégia:
        - Tenta usar o endpoint timeseries do exchangerate.host para cada mês.
        - Em caso de falha (ex: 5xx), faz até N tentativas com backoff exponencial.
        - Se continuar falhando, tenta fallback com uma requisição por mês ao serviço frankfurter.app usando a data central do mês (15º dia) como proxy da média.
        - Registra avisos no log de validação quando não for possível obter taxas e preenche com None para esses meses.
        """
        moedas = list(set(moedas))
        cache_key = (ano, mes_final, tuple(sorted(moedas)))
        if cache_key in self.cache_cambio:
            return self.cache_cambio[cache_key]

        resultado = {m: {} for m in moedas}

        if requests is None:
            self._log_validacao('AVISO', 'Biblioteca requests não disponível; taxas de câmbio não serão buscadas.', {'moedas': moedas})
            self.cache_cambio[cache_key] = resultado
            return resultado

        def _fetch_with_retries(url, params=None, headers=None, max_retries=3, backoff=1.0):
            last_exc = None
            for attempt in range(1, max_retries + 1):
                try:
                    if requests is None:
                        raise Exception('requests não disponível')
                    r = requests.get(url, params=params, headers=headers, timeout=12)
                    if r.status_code == 200:
                        return r
                    last_exc = Exception(f"Erro na API: {r.status_code}")
                except Exception as e:
                    last_exc = e
                # backoff
                time.sleep(backoff * (2 ** (attempt - 1)))
            # Após tentativas, relança a última exceção encapsulada
            if last_exc is None:
                raise Exception('Falha desconhecida na requisição')
            raise last_exc

        for moeda in moedas:
            for mes in range(1, mes_final + 1):
                try:
                    primeiro_dia = datetime(ano, mes, 1).date()
                    ultimo_dia = datetime(ano, mes, calendar.monthrange(ano, mes)[1]).date()
                    url_timeseries = f"https://api.exchangerate.host/timeseries"
                    params = {'start_date': primeiro_dia.isoformat(), 'end_date': ultimo_dia.isoformat(), 'base': 'BRL', 'symbols': moeda}
                    resp = None
                    try:
                        r = _fetch_with_retries(url_timeseries, params=params, max_retries=3, backoff=1.0)
                        resp = r
                    except Exception as e_ts:
                        # registrar aviso e tentar fallback por dia central do mês
                        self._log_validacao('AVISO', f'Timeseries falhou para {moeda} mês {mes}: {e_ts}. Tentando fallback diário.', {'moeda': moeda, 'mes': mes})

                    rates = []
                    if resp is not None:
                        data = resp.json()
                        for d, vals in data.get('rates', {}).items():
                            v = vals.get(moeda)
                            if v is not None:
                                rates.append(v)

                    if rates:
                        resultado[moeda][mes] = sum(rates) / len(rates)
                        continue

                    # Fallback: usar data central do mês (15º dia ou último dia se menor)
                    dia_central = min(15, calendar.monthrange(ano, mes)[1])
                    data_central = datetime(ano, mes, dia_central).date().isoformat()
                    # Primeiro tentar frankfurter
                    try:
                        url_fk = f"https://api.frankfurter.app/{data_central}"
                        params_fk = {'from': 'BRL', 'to': moeda}
                        r_fk = _fetch_with_retries(url_fk, params=params_fk, max_retries=3, backoff=1.0)
                        j_fk = r_fk.json()
                        v_fk = j_fk.get('rates', {}).get(moeda)
                        if v_fk is not None:
                            resultado[moeda][mes] = v_fk
                            continue
                    except Exception as e_fk:
                        self._log_validacao('AVISO', f'Fallback frankfurter falhou para {moeda} mês {mes}: {e_fk}', {'moeda': moeda, 'mes': mes})

                    # Último recurso: tentar endpoint convert do exchangerate.host para a data_central
                    try:
                        url_conv = f"https://api.exchangerate.host/convert"
                        params_conv = {'from': 'BRL', 'to': moeda, 'date': data_central}
                        r_conv = _fetch_with_retries(url_conv, params=params_conv, max_retries=2, backoff=1.0)
                        j_conv = r_conv.json()
                        # resultado pode vir em 'result' ou em 'info' dependendo do endpoint
                        v_conv = j_conv.get('result') if 'result' in j_conv else j_conv.get('info', {}).get('rate')
                        if v_conv is not None:
                            resultado[moeda][mes] = float(v_conv)
                            continue
                    except Exception as e_conv:
                        self._log_validacao('AVISO', f'Fallback convert falhou para {moeda} mês {mes}: {e_conv}', {'moeda': moeda, 'mes': mes})

                    # Se tudo falhar, preenche com None e registra
                    resultado[moeda][mes] = None
                    self._log_validacao('AVISO', f'Não foi possível obter taxa para {moeda} mês {mes}; valor ficará vazio.', {'moeda': moeda, 'mes': mes})

                except Exception as e_outer:
                    resultado[moeda][mes] = None
                    self._log_validacao('AVISO', f'Erro inesperado ao obter taxa para {moeda} mês {mes}: {e_outer}', {'moeda': moeda, 'mes': mes})

        self.cache_cambio[cache_key] = resultado
        return resultado

    # ------------------ Estado e Reconciliacao (Recebimentos) ------------------
    def _carregar_estado(self):
        """Carrega ou inicializa o arquivo de estado que guarda adiantamentos e reconciliações."""
        try:
            # Definir esquema correto do estado
            expected = ['PROCESSO', 'VALOR_TOTAL_PROCESSO', 'TOTAL_PAGO_ACUMULADO', 'TOTAL_ADIANTADO_COMISSAO',
                        'STATUS_PAGAMENTO', 'STATUS_RECONCILIACAO', 'STATUS_PROCESSO_ANALISE', 'ULTIMA_ATUALIZACAO']
            if os.path.exists(ARQUIVO_ESTADO):
                try:
                    df_estado = pd.read_excel(ARQUIVO_ESTADO, sheet_name='ESTADO')
                except Exception:
                    df_estado = pd.read_excel(ARQUIVO_ESTADO)
                # Garantir colunas e tipos
                for c in expected:
                    if c not in df_estado.columns:
                        df_estado[c] = None
                # Normalizar colunas e preencher NaNs
                df_estado = df_estado[expected].copy()
                df_estado['TOTAL_PAGO_ACUMULADO'] = pd.to_numeric(df_estado['TOTAL_PAGO_ACUMULADO'], errors='coerce').fillna(0.0)
                df_estado['TOTAL_ADIANTADO_COMISSAO'] = pd.to_numeric(df_estado['TOTAL_ADIANTADO_COMISSAO'], errors='coerce').fillna(0.0)
                df_estado['VALOR_TOTAL_PROCESSO'] = pd.to_numeric(df_estado['VALOR_TOTAL_PROCESSO'], errors='coerce').fillna(0.0)
                if 'STATUS_PAGAMENTO' in df_estado.columns:
                    df_estado['STATUS_PAGAMENTO'] = df_estado['STATUS_PAGAMENTO'].astype(str).replace({'nan': None})
                self.estado = df_estado
            else:
                # Iniciar DataFrame vazio com as colunas corretas
                self.estado = pd.DataFrame(columns=expected)
        except Exception as e:
            self._log_validacao('AVISO', f'Falha ao carregar estado ({ARQUIVO_ESTADO}): {e}', {})
            # Garantir esquema mínimo em caso de erro
            self.estado = pd.DataFrame(columns=['PROCESSO', 'VALOR_TOTAL_PROCESSO', 'TOTAL_PAGO_ACUMULADO', 'TOTAL_ADIANTADO_COMISSAO', 'STATUS_PAGAMENTO', 'STATUS_RECONCILIACAO', 'STATUS_PROCESSO_ANALISE', 'ULTIMA_ATUALIZACAO'])

    def _salvar_estado(self):
        """Salva o dataframe de estado no arquivo ARQUIVO_ESTADO."""
        try:
            if getattr(self, 'estado', None) is None:
                return
            with pd.ExcelWriter(ARQUIVO_ESTADO, engine='openpyxl') as w:
                # garantir tipos simples
                df = self.estado.copy()
                df.to_excel(w, sheet_name='ESTADO', index=False)
            if getattr(self, '_logger', None):
                self._logger.info(f"Estado salvo em {ARQUIVO_ESTADO}")
        except Exception as e:
            self._log_validacao('AVISO', f'Falha ao salvar estado em {ARQUIVO_ESTADO}: {e}', {})

    def _get_valor_total_processo(self, proc):
        """Retorna a soma de 'Valor Realizado' de todos os itens do processo no arquivo ANALISE_COMERCIAL_COMPLETA.

        Proc pode ser string ou número; fazemos comparação por string trimmed.
        Retorna float (0.0 se não encontrado ou erro).
        """
        try:
            df_anal = self.data.get('ANALISE_COMERCIAL_COMPLETA', pd.DataFrame())
            if df_anal.empty:
                return 0.0
            # Normalizar coluna de processo
            possible_proc_cols = [c for c in df_anal.columns if str(c).strip().lower() == 'processo']
            if not possible_proc_cols:
                # tentar variações
                for cand in ('PROCESSO','processo'):
                    if cand in df_anal.columns:
                        possible_proc_cols = [cand]
                        break
            if not possible_proc_cols:
                return 0.0
            proc_col = possible_proc_cols[0]
            proc_s = str(proc).strip()
            mask = df_anal[proc_col].astype(str).str.strip() == proc_s
            subset = df_anal[mask]
            if subset.empty:
                # tentar correspondência numérica
                try:
                    proc_int = int(float(proc))
                    mask2 = df_anal[proc_col].astype(str).str.strip() == str(proc_int)
                    subset = df_anal[mask2]
                except Exception:
                    pass
            if subset.empty:
                return 0.0
            # possíveis colunas de valor realizado
            valor_cols = [c for c in subset.columns if str(c).strip().lower() in ('valor realizado','valor_realizado','valorrealizado','valor realizado total','valor realizado (brl)')]
            if not valor_cols:
                # tentar nomes com acentos/alternativas
                alt = [c for c in subset.columns if 'valor' in str(c).lower() and 'real' in str(c).lower()]
                valor_cols = alt
            if not valor_cols:
                return 0.0
            # somar valores numéricos
            total = 0.0
            for c in valor_cols:
                try:
                    total += float(pd.to_numeric(subset[c], errors='coerce').fillna(0.0).sum())
                except Exception:
                    continue
            return float(total)
        except Exception:
            return 0.0

    def _aplicar_adiantamentos_recebimentos(self):
        """Calcula e aplica adiantamentos de comissão baseados nos recebimentos do mês.

        Estratégia:
        - Para cada PROCESSO presente em RECEBIMENTOS, soma a comissão calculada (original) para o processo.
        - Calcula um adiantamento percentual (parâmetro 'percentual_adiantamento_recebimento', default=0.5)
          sobre a comissão total e cria linhas adicionais em self.comissoes_df com tipo_lancamento='Adiantamento Recebimento'.
        - Atualiza self.estado.TOTAL_ADIANTADO_COMISSAO para cada processo.
        """
        try:
            df_rec = self.data.get('RECEBIMENTOS', pd.DataFrame())
            if df_rec.empty:
                return

            if not hasattr(self, 'comissoes_df') or self.comissoes_df.empty:
                return

            novas_linhas = []
            df_fat = self.data.get('FATURADOS', pd.DataFrame())
            df_atr = self.data.get('ATRIBUICOES', pd.DataFrame())
            df_colabs_com_cargos = self.data.get('COLABORADORES', pd.DataFrame())
            df_status_pag = self.data.get('STATUS_PAGAMENTOS', pd.DataFrame())

            def _normalize_proc(val):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    return None
                if isinstance(val, (int, np.integer)):
                    return str(int(val))
                if isinstance(val, float):
                    if pd.isna(val):
                        return None
                    if float(val).is_integer():
                        return str(int(val))
                    return str(val).strip()
                s = str(val).strip()
                if not s:
                    return None
                try:
                    s_float = float(s.replace(',', '.'))
                    if float(s_float).is_integer():
                        return str(int(s_float))
                except Exception:
                    pass
                return s

            def _find_column(df, aliases):
                for col in df.columns:
                    if str(col).strip().lower() in aliases:
                        return col
                return None

            status_map = {}
            valor_map = {}
            if df_status_pag is not None and not df_status_pag.empty:
                proc_col = _find_column(df_status_pag, {'processo', 'id processo', 'id_processo'})
                status_col = _find_column(df_status_pag, {'status_pagamento', 'status pagamento', 'status do pagamento', 'status_pagamento_processo', 'status'})
                valor_col = _find_column(df_status_pag, {'valor_original', 'valor original', 'valor', 'valor do processo'})

                if proc_col is not None:
                    for _, row in df_status_pag.iterrows():
                        key = _normalize_proc(row.get(proc_col))
                        if not key:
                            continue
                        if status_col is not None:
                            status_map[key] = row.get(status_col)
                        if valor_col is not None:
                            existing = valor_map.get(key)
                            if existing is None or pd.isna(existing):
                                valor_map[key] = row.get(valor_col)

            cargos_gestao = df_colabs_com_cargos[df_colabs_com_cargos['tipo_cargo'] == 'Gestão']['cargo'].unique() if not df_colabs_com_cargos.empty else []
            df_atribuicoes_gestao = df_atr[df_atr['cargo'].isin(cargos_gestao)] if not df_atr.empty else pd.DataFrame()

            # Processar cada recebimento: atualizar estado (criar nova linha se necessário)
            for _, rec in df_rec.iterrows():
                proc = rec.get('PROCESSO')
                valor_recebido = rec.get('VALOR_RECEBIDO')
                if pd.isna(proc) or pd.isna(valor_recebido):
                    continue

                proc = int(proc) if (not pd.isna(proc) and str(proc).strip().isdigit()) else str(proc).strip()
                proc_norm = _normalize_proc(proc)

                # Obter VALOR_ORIGINAL do arquivo Status_Pagamentos_Processos, se disponível
                valor_original = valor_map.get(proc_norm)
                status_pag_val = status_map.get(proc_norm)
                try:
                    if not df_status_pag.empty and 'PROCESSO' in df_status_pag.columns:
                        sp = df_status_pag[df_status_pag['PROCESSO'] == proc]
                        if sp.empty and 'Processo' in df_status_pag.columns:
                            sp = df_status_pag[df_status_pag['Processo'] == proc]
                        if not sp.empty:
                            # procurar colunas possíveis para VALOR_ORIGINAL
                            for col_candidate in ('VALOR_ORIGINAL', 'Valor Original', 'Valor_Original', 'VALOR'):
                                if col_candidate in sp.columns:
                                    valor_original = sp.iloc[0][col_candidate]
                                    break
                            if status_pag_val is None:
                                for status_candidate in ('STATUS_PAGAMENTO', 'Status Pagamento', 'Status_Pagamento', 'STATUS', 'Status'):
                                    if status_candidate in sp.columns:
                                        status_pag_val = sp.iloc[0][status_candidate]
                                        break
                except Exception:
                    valor_original = None
                # Preferir valor do Analise_Comercial_Completa (Valor Orçado e aliases) quando existir
                try:
                    df_analise_local2 = self.data.get('ANALISE_COMERCIAL_COMPLETA', pd.DataFrame())
                    if not df_analise_local2.empty and 'Processo' in df_analise_local2.columns:
                        sp_a2 = df_analise_local2[df_analise_local2['Processo'].astype(str).str.strip() == str(proc)]
                        if not sp_a2.empty:
                            aliases = [
                                'Valor Orçado','Valor Orcado','Valor Orçado Total','Valor Orcado Total',
                                'Valor do Orçado','Valor do Orcado'
                            ]
                            for alt in aliases:
                                if alt in sp_a2.columns:
                                    cand = sp_a2.iloc[0][alt]
                                    if cand is not None and not pd.isna(cand):
                                        valor_original = cand
                                        break
                except Exception:
                    pass

                # Se o processo não existir no estado, criar uma nova linha
                sidx = self.estado[self.estado['PROCESSO'] == proc].index
                if len(sidx) == 0:
                    # priorizar soma dos itens do processo no arquivo de análise comercial
                    try:
                        total_proc = self._get_valor_total_processo(proc)
                        if total_proc and total_proc > 0:
                            vtp_val = float(total_proc)
                        else:
                            vtp_val = float(valor_original) if valor_original is not None and not pd.isna(valor_original) else 0.0
                    except Exception:
                        vtp_val = float(valor_original) if valor_original is not None and not pd.isna(valor_original) else 0.0

                    nova = {
                        'PROCESSO': proc,
                        'VALOR_TOTAL_PROCESSO': vtp_val,
                        'TOTAL_PAGO_ACUMULADO': float(valor_recebido),
                        'TOTAL_ADIANTADO_COMISSAO': 0.0,
                        'STATUS_PAGAMENTO': status_pag_val,
                        'STATUS_RECONCILIACAO': 'Nao Realizada',
                        'STATUS_PROCESSO_ANALISE': None,
                        'ULTIMA_ATUALIZACAO': datetime.now().isoformat()
                    }
                    self.estado = pd.concat([self.estado, pd.DataFrame([nova])], ignore_index=True, sort=False)
                else:
                    idx0 = sidx[0]
                    # somar valor recebido
                    try:
                        prev_pago = pd.to_numeric(self.estado.at[idx0, 'TOTAL_PAGO_ACUMULADO'], errors='coerce')
                        prev_pago = float(prev_pago) if not pd.isna(prev_pago) else 0.0
                    except Exception:
                        prev_pago = 0.0
                    try:
                        vrec = float(valor_recebido)
                    except Exception:
                        vrec = pd.to_numeric(valor_recebido, errors='coerce')
                        vrec = float(vrec) if not pd.isna(vrec) else 0.0
                    self.estado.at[idx0, 'TOTAL_PAGO_ACUMULADO'] = prev_pago + vrec
                    if status_pag_val is not None and not (isinstance(status_pag_val, float) and pd.isna(status_pag_val)):
                        self.estado.at[idx0, 'STATUS_PAGAMENTO'] = status_pag_val
                    # atualizar VALOR_TOTAL_PROCESSO se estiver vazio
                    try:
                        vtp = pd.to_numeric(self.estado.at[idx0, 'VALOR_TOTAL_PROCESSO'], errors='coerce')
                        vtp_val = float(vtp) if not pd.isna(vtp) else 0.0
                        if vtp is None or pd.isna(vtp) or vtp_val == 0.0:
                            if valor_original is not None and not pd.isna(valor_original):
                                try:
                                    self.estado.at[idx0, 'VALOR_TOTAL_PROCESSO'] = float(pd.to_numeric(valor_original, errors='coerce'))
                                except Exception:
                                    self.estado.at[idx0, 'VALOR_TOTAL_PROCESSO'] = 0.0
                    except Exception:
                        pass
                    # atualizar timestamp
                    self.estado.at[idx0, 'ULTIMA_ATUALIZACAO'] = datetime.now().isoformat()

                # Nota: nesta etapa apenas registramos/atualizamos o estado com os recebimentos.
                # A geração de linhas de adiantamento (novas_linhas) permanece separada e só
                # será executada se houver lógica adicional preenchendo `novas_linhas`.

            # Após processar todos os recebimentos, atualizar STATUS_PROCESSO_ANALISE
            try:
                df_analise = self.data.get('ANALISE_COMERCIAL_COMPLETA', pd.DataFrame())
                if not df_analise.empty and 'Processo' in df_analise.columns and 'Status Processo' in df_analise.columns:
                    # construir mapa processo -> status (string)
                    mapa_status = df_analise.set_index(df_analise['Processo'].astype(str).str.strip())['Status Processo'].to_dict()
                else:
                    mapa_status = {}
            except Exception:
                mapa_status = {}

            for idx in self.estado.index:
                try:
                    proc_key = str(self.estado.at[idx, 'PROCESSO']).strip()
                    status_proc = mapa_status.get(proc_key)
                    if status_proc is None:
                        # tentar procurar por inteiros/sem formatação
                        status_proc = mapa_status.get(str(int(float(proc_key))) if proc_key.replace('.','',1).isdigit() else None)
                    self.estado.at[idx, 'STATUS_PROCESSO_ANALISE'] = status_proc
                except Exception:
                    # não bloquear o fluxo se houver problemas de formatação
                    self.estado.at[idx, 'STATUS_PROCESSO_ANALISE'] = None
 
            if status_map:
                for idx in self.estado.index:
                    try:
                        proc_norm_all = _normalize_proc(self.estado.at[idx, 'PROCESSO'])
                        if proc_norm_all and proc_norm_all in status_map:
                            status_val = status_map.get(proc_norm_all)
                            if status_val is not None and not (isinstance(status_val, float) and pd.isna(status_val)):
                                self.estado.at[idx, 'STATUS_PAGAMENTO'] = status_val
                    except Exception:
                        continue

            if novas_linhas:
                try:
                    df_novas = pd.DataFrame(novas_linhas)
                    self.comissoes_df = pd.concat([self.comissoes_df, df_novas], ignore_index=True, sort=False)
                    # Atualizar TOTAL_ADIANTADO_COMISSAO no estado para cada processo afetado
                    for _, r in df_novas.iterrows():
                        p = r.get('processo')
                        val = r.get('comissao_calculada', 0.0)
                        sidx = self.estado[self.estado['PROCESSO'] == p].index
                        norm_p = _normalize_proc(p)
                        status_retro = status_map.get(norm_p)
                        if len(sidx) == 0:
                            nova = {
                                'PROCESSO': p,
                                'VALOR_TOTAL_PROCESSO': 0.0,
                                'TOTAL_PAGO_ACUMULADO': 0.0,
                                'TOTAL_ADIANTADO_COMISSAO': float(val) if not pd.isna(val) else 0.0,
                                'STATUS_PAGAMENTO': status_retro,
                                'STATUS_RECONCILIACAO': 'Nao Realizada',
                                'STATUS_PROCESSO_ANALISE': None,
                                'ULTIMA_ATUALIZACAO': datetime.now().isoformat()
                            }
                            self.estado = pd.concat([self.estado, pd.DataFrame([nova])], ignore_index=True, sort=False)
                        else:
                            idx0 = sidx[0]
                            prev = pd.to_numeric(self.estado.at[idx0, 'TOTAL_ADIANTADO_COMISSAO'], errors='coerce')
                            prev = float(prev) if not pd.isna(prev) else 0.0
                            val_num = pd.to_numeric(val, errors='coerce')
                            val_num = float(val_num) if not pd.isna(val_num) else 0.0
                            self.estado.at[idx0, 'TOTAL_ADIANTADO_COMISSAO'] = prev + val_num
                            self.estado.at[idx0, 'ULTIMA_ATUALIZACAO'] = datetime.now().isoformat()
                            if status_retro is not None and not (isinstance(status_retro, float) and pd.isna(status_retro)):
                                self.estado.at[idx0, 'STATUS_PAGAMENTO'] = status_retro
                except Exception as e:
                    self._log_validacao('AVISO', f'Falha ao anexar linhas de adiantamento: {e}', {})

        except Exception as e:
            self._log_validacao('AVISO', f'Erro ao aplicar adiantamentos de recebimentos: {e}', {})

        # --- Gerar COMISSOES_RECEBIMENTO separada para os colaboradores que recebem por recebimento ---
        try:
            self.comissoes_recebimento_df = pd.DataFrame()
            df_rec = self.data.get('RECEBIMENTOS', pd.DataFrame())
            df_analise = self.data.get('ANALISE_COMERCIAL_COMPLETA', pd.DataFrame())
            df_colabs = self.data.get('COLABORADORES', pd.DataFrame())
            df_fat = self.data.get('FATURADOS', pd.DataFrame())

            # Exigir ANALISE_COMERCIAL_COMPLETA: sem fallback permitido
            if df_analise is None or df_analise.empty:
                self._log_validacao('ERRO', 'ANALISE_COMERCIAL_COMPLETA ausente; não é permitido fallback para FATURADOS. COMISSOES_RECEBIMENTO não será gerada.', {})
                return
            df_map = df_analise
            map_source = 'ANALISE_COMERCIAL_COMPLETA'

            if not df_rec.empty and self.recebe_por_recebimento:
                rows = []
                total_rec = 0
                total_matched = 0
                total_unmatched = 0
                # helper: tenta mapear um recebimento usando apenas a tabela ANALISE_COMERCIAL_COMPLETA
                def _map_recebimento(proc_val, valor_val, id_cliente_val, df_map_local):
                    # Normalizar processo para comparar números equivalentes (ex.: 999999 vs 999999.0)
                    def _norm_proc(v):
                        try:
                            s = str(v).strip()
                            # se for numerico-like, converter para int para remover .0
                            if s.replace('.', '', 1).isdigit():
                                try:
                                    if '.' in s:
                                        f = float(s)
                                        i = int(f)
                                        if f == float(i):
                                            return str(i)
                                except Exception:
                                    pass
                                # número inteiro puro como string
                                return str(int(float(s)))
                            return s
                        except Exception:
                            return str(v).strip()

                    proc_s = _norm_proc(proc_val)
                    # 1) exact match (normalizado)
                    if 'Processo' in df_map_local.columns:
                        map_proc_norm = df_map_local['Processo'].apply(_norm_proc)
                        exact_idx = map_proc_norm == proc_s
                        if exact_idx.any():
                            return df_map_local[exact_idx].iloc[0], 'exact_map'

                    # 2) substring match (normalizado)
                    if 'Processo' in df_map_local.columns:
                        map_proc_norm = df_map_local['Processo'].apply(_norm_proc)
                        mask_sub = map_proc_norm.apply(lambda x: (x in proc_s) or (proc_s in x))
                        cand = df_map_local[mask_sub]
                        if not cand.empty:
                            # choose candidate closest by amount when possible
                            if 'Valor Realizado' in cand.columns and valor_val is not None:
                                cand = cand.copy()
                                try:
                                    cand['diff'] = cand['Valor Realizado'].apply(lambda x: abs((float(x) if pd.notna(x) else 0.0) - float(valor_val)))
                                    cand_sorted = cand.sort_values('diff')
                                    return cand_sorted.iloc[0], 'substring_amount_best'
                                except Exception:
                                    pass
                            return cand.iloc[0], 'substring_first'

                    # 3) no match
                    if getattr(self, '_logger', None):
                        try:
                            self._logger.info(f"[Receb] Sem mapeamento para processo {proc_val} (normalizado='{proc_s}') em {map_source}.")
                        except Exception:
                            pass
                    return None, None

                    # 3) match by client + approximate amount
                    try:
                        if id_cliente_val is not None and 'Cliente' in df_fat_local.columns:
                            same_cli = df_fat_local[df_fat_local['Cliente'] == id_cliente_val]
                            if not same_cli.empty and 'Valor Realizado' in same_cli.columns:
                                same_cli['diff'] = same_cli['Valor Realizado'].apply(lambda x: abs((float(x) if pd.notna(x) else 0.0) - float(valor_val)))
                                cand_sorted = same_cli.sort_values('diff').copy()
                                if cand_sorted.iloc[0]['diff'] <= max(1.0, 0.01 * float(valor_val)):
                                    return cand_sorted.iloc[0], 'client_amount'
                    except Exception:
                        pass

                    # 4) numeric prefix reduction: try removing last k digits from proc and compare
                    try:
                        pnum = int(proc_s)
                        for k in range(1, 5):
                            truncated = str(pnum // (10 ** k))
                            cand2 = df_fat_local[df_fat_local['Processo'].astype(str).str.strip() == truncated]
                            if not cand2.empty:
                                return cand2.iloc[0], f'truncate_{k}'
                    except Exception:
                        pass

                    return None, 'no_match'

                for _, rec in df_rec.iterrows():
                    total_rec += 1
                    proc = rec.get('PROCESSO')
                    valor_recebido = rec.get('VALOR_RECEBIDO')
                    if pd.isna(proc) or pd.isna(valor_recebido):
                        continue
                    match_row, why = _map_recebimento(proc, valor_recebido, rec.get('ID_CLIENTE', None), df_map)
                    if match_row is None:
                        total_unmatched += 1
                        placeholder = {
                            'id_colaborador': None,
                            'nome_colaborador': None,
                            'cargo': None,
                            'processo': proc,
                            'linha': None, 'grupo': None, 'subgrupo': None, 'tipo_mercadoria': None,
                            'faturamento_item': valor_recebido,
                            'taxa_rateio_aplicada': None,
                            'percentual_elegibilidade_pe': None,
                            'fator_correcao_fc': None,
                            'comissao_calculada': None,
                            'tipo_lancamento': 'Recebimento',
                            'observacao': f'Processo não mapeado em {map_source}',
                            'mapping_found': False
                        }
                        rows.append(placeholder)
                        continue
                    total_matched += 1
                    primeira = match_row
                    if getattr(self, '_logger', None):
                        self._logger.info(f"Processo {proc} mapeado via {why} para processo faturado {primeira.get('Processo')}")
                    contexto = {
                        'linha': primeira.get('Negócio'), 'grupo': primeira.get('Grupo'),
                        'subgrupo': primeira.get('Subgrupo'), 'tipo_mercadoria': primeira.get('Tipo de Mercadoria')
                    }

                    # identificar colaboradores responsáveis (gestão e operacional) usando as mesmas regras de atribuições
                    df_atr = self.data.get('ATRIBUICOES', pd.DataFrame())
                    nomes_operacionais = []
                    if pd.notna(primeira.get('Consultor Interno')):
                        nomes_operacionais.append(primeira.get('Consultor Interno'))
                    if pd.notna(primeira.get('Representante-pedido')):
                        nomes_operacionais.append(primeira.get('Representante-pedido'))

                    cargos_gestao = df_colabs[df_colabs['tipo_cargo'] == 'Gestão']['cargo'].unique() if not df_colabs.empty else []
                    df_atribuicoes_gestao = df_atr[df_atr['cargo'].isin(cargos_gestao)] if not df_atr.empty else pd.DataFrame()

                    atribuidos_gestao = df_atribuicoes_gestao[
                        (df_atribuicoes_gestao['linha'] == contexto['linha']) &
                        (df_atribuicoes_gestao['grupo'] == contexto['grupo']) &
                        (df_atribuicoes_gestao['subgrupo'] == contexto['subgrupo']) &
                        (df_atribuicoes_gestao['tipo_mercadoria'] == contexto['tipo_mercadoria'])
                    ] if not df_atribuicoes_gestao.empty else pd.DataFrame(columns=['colaborador','cargo'])

                    atribuidos_operacional = df_colabs[df_colabs['nome_colaborador'].isin(nomes_operacionais)] if not df_colabs.empty else pd.DataFrame(columns=['nome_colaborador','cargo'])

                    colaboradores_para_comissionar = pd.concat([
                        atribuidos_gestao[['colaborador', 'cargo']] if not atribuidos_gestao.empty else pd.DataFrame(columns=['colaborador','cargo']),
                        atribuidos_operacional[['nome_colaborador', 'cargo']].rename(columns={'nome_colaborador': 'colaborador'}) if not atribuidos_operacional.empty else pd.DataFrame(columns=['colaborador','cargo'])
                    ]).drop_duplicates().reset_index(drop=True)

                    # Filtrar SOMENTE para colaboradores que recebem por recebimento
                    colaboradores_receb = [c for c in colaboradores_para_comissionar['colaborador'].tolist() if c in self.recebe_por_recebimento]
                    if getattr(self, '_logger', None):
                        self._logger.info(f"Processo {proc}: colaboradores para recebimento identificados: {colaboradores_receb}")
                    if not colaboradores_receb:
                        continue
                    for colab in colaboradores_receb:
                        row_col = df_colabs[df_colabs['nome_colaborador'] == colab]
                        cargo = row_col.iloc[0]['cargo'] if not row_col.empty else None
                        regra = self._get_regra_comissao(contexto['linha'], contexto['grupo'], contexto['subgrupo'], contexto['tipo_mercadoria'], cargo)
                        if regra is None:
                            continue
                        taxa_rateio = regra['taxa_rateio_maximo_pct'] / 100.0
                        pe = regra['fatia_cargo_pct'] / 100.0
                        com_calc = float(valor_recebido) * taxa_rateio * pe
                        linha_receb = {
                            'id_colaborador': row_col.iloc[0]['id_colaborador'] if not row_col.empty else None,
                            'nome_colaborador': colab,
                            'cargo': cargo,
                            'processo': proc,
                            'linha': contexto['linha'], 'grupo': contexto['grupo'], 'subgrupo': contexto['subgrupo'], 'tipo_mercadoria': contexto['tipo_mercadoria'],
                            'faturamento_item': valor_recebido,
                            'taxa_rateio_aplicada': taxa_rateio,
                            'percentual_elegibilidade_pe': pe,
                            'fator_correcao_fc': 1.0,
                            'comissao_calculada': com_calc,
                            'tipo_lancamento': 'Recebimento',
                            'observacao': 'Comissão por Recebimento'
                        }
                        if getattr(self, '_logger', None):
                            self._logger.info(f"Linha gerada para COMISSOES_RECEBIMENTO: {linha_receb}")
                        rows.append(linha_receb)
                # Ao final, criar DataFrame mesmo se só placeholders
                self.comissoes_recebimento_df = pd.DataFrame(rows) if rows else pd.DataFrame()
                if getattr(self, '_logger', None):
                    self._logger.info(f"Recebimentos processados: total={total_rec}, matched={total_matched}, unmatched={total_unmatched}, linhas geradas={len(rows)}")
                # Registrar na validação para inspeção
                self._log_validacao('INFO', f'Recebimentos processados: total={total_rec}, matched={total_matched}, unmatched={total_unmatched}', {'total': total_rec, 'matched': total_matched, 'unmatched': total_unmatched})
        except Exception as e:
            self._log_validacao('AVISO', f'Erro ao gerar COMISSOES_RECEBIMENTO: {e}', {})

        # Atualizar TOTAL_ADIANTADO_COMISSAO acumulando a partir de COMISSOES_RECEBIMENTO
        try:
            if hasattr(self, 'comissoes_recebimento_df') and self.comissoes_recebimento_df is not None and not self.comissoes_recebimento_df.empty:
                soma_proc = self.comissoes_recebimento_df.groupby('processo', dropna=False)['comissao_calculada'].sum().to_dict()
                for i in self.estado.index:
                    p = self.estado.at[i, 'PROCESSO']
                    if p in soma_proc:
                        try:
                            prev = pd.to_numeric(self.estado.at[i, 'TOTAL_ADIANTADO_COMISSAO'], errors='coerce')
                            prev = float(prev) if not pd.isna(prev) else 0.0
                            add = float(soma_proc[p]) if not pd.isna(soma_proc[p]) else 0.0
                            self.estado.at[i, 'TOTAL_ADIANTADO_COMISSAO'] = prev + add
                        except Exception:
                            pass
        except Exception as e:
            self._log_validacao('AVISO', f'Falha ao atualizar TOTAL_ADIANTADO_COMISSAO a partir de COMISSOES_RECEBIMENTO: {e}', {})


    def _executar_reconciliacoes(self):
        _info("Iniciando reconciliações de comissões por recebimento...")
        self.reconciliacao_detalhada_list = []
        self.reconciliacao_resumo_list = []

        try:
            self._sync_estado_from_inputs()
        except Exception:
            pass
        estado_df = getattr(self, 'estado', None)
        if estado_df is None or estado_df.empty:
            _info("  Nenhum processo elegível no estado. Reconciliações não executadas.")
            return
        for idx, row in estado_df.iterrows():
            proc = row.get('PROCESSO')
            if proc is None:
                continue
            if isinstance(proc, float) and pd.isna(proc):
                continue
            proc_str = str(proc).strip()
            if not proc_str:
                continue

            try:
                status_pagamento = _normalize_text(row.get('STATUS_PAGAMENTO', ''))
                status_analise = _normalize_text(row.get('STATUS_PROCESSO_ANALISE', ''))
                status_reconc = _normalize_text(row.get('STATUS_RECONCILIACAO', ''))

                is_quitado = 'QUITADO' in status_pagamento
                is_faturado = (status_analise == 'FATURADO')

                if status_reconc in ('REALIZADA', 'CONCLUIDA'):
                    continue
                if not (is_quitado and is_faturado):
                    continue
            except Exception as e:
                try:
                    self.estado.loc[idx, 'STATUS_RECONCILIACAO'] = f'Erro: {e}'
                except Exception:
                    pass
                continue

            try:
                total_adiantado = pd.to_numeric(row.get('TOTAL_ADIANTADO_COMISSAO', 0.0), errors='coerce')
            except Exception:
                total_adiantado = 0.0
            total_adiantado = float(total_adiantado) if not pd.isna(total_adiantado) else 0.0

            try:
                linhas_detalhadas, comissao_correta_total = self._gerar_reconciliacao_detalhada_processo(proc_str)
                if not linhas_detalhadas:
                    _info(f"  [AVISO] Reconciliação para {proc_str} não gerou linhas. Pulando.")
                    continue

                saldo_final = comissao_correta_total - total_adiantado
                self.reconciliacao_detalhada_list.extend(linhas_detalhadas)
                self.reconciliacao_resumo_list.append({
                    'PROCESSO': proc_str,
                    'COMISSAO_CORRETA_TOTAL': comissao_correta_total,
                    'TOTAL_ADIANTAMENTOS_PAGOS': total_adiantado,
                    'SALDO_FINAL_RECONCILIACAO': saldo_final
                })
                try:
                    self.estado.loc[idx, 'STATUS_RECONCILIACAO'] = 'Realizada'
                except Exception:
                    pass
                _info(f"  [SUCESSO] Reconciliação para {proc_str} concluída. Saldo: {saldo_final:.2f}")
            except Exception as e:
                _info(f"  [ERRO] Falha ao processar reconciliação para {proc_str}: {e}")
                try:
                    self.estado.loc[idx, 'STATUS_RECONCILIACAO'] = f'Erro: {e}'
                except Exception:
                    pass


    def _gerar_reconciliacao_detalhada_processo(self, proc_id):
        """Calcula a comissão correta retroativa para um processo, item a item,
        reprocessando item a item com os dados de performance históricos do mês de faturamento.
        """
        def _match_column(df, candidates):
            if df is None or df.empty:
                return None
            normalized = { _normalize_text(col): col for col in df.columns }
            for cand in candidates:
                cand_norm = _normalize_text(cand)
                if cand_norm in normalized:
                    return normalized[cand_norm]
            return None

        def _safe_float(value):
            try:
                series_val = pd.to_numeric(pd.Series([value]), errors='coerce')
                result = float(series_val.iloc[0])
            except Exception:
                try:
                    result = float(value)
                except Exception:
                    result = 0.0
            return 0.0 if pd.isna(result) else float(result)

        def _build_group_series(df, group_col, value_col):
            if df is None or df.empty or group_col is None or value_col is None:
                return pd.Series(dtype=float)
            data = df[[group_col, value_col]].copy()
            data[group_col] = data[group_col].astype(str).str.strip()
            data[value_col] = pd.to_numeric(data[value_col], errors='coerce').fillna(0.0)
            return data.groupby(group_col)[value_col].sum()

        def _build_rentabilidade_series(df):
            if df is None or df.empty:
                return pd.Series(dtype=float)
            linha_col = _match_column(df, ['negocio', 'negócio', 'linha'])
            grupo_col = _match_column(df, ['grupo'])
            subgrupo_col = _match_column(df, ['subgrupo'])
            tipo_col = _match_column(df, ['tipo de mercadoria', 'tipo mercadoria'])
            valor_col = _match_column(df, ['rentabilidade_realizada_pct', 'rentabilidade', 'rentabilidade realizada'])
            if None in (linha_col, grupo_col, subgrupo_col, tipo_col, valor_col):
                return pd.Series(dtype=float)
            df_copy = df[[linha_col, grupo_col, subgrupo_col, tipo_col, valor_col]].copy()
            for col in (linha_col, grupo_col, subgrupo_col, tipo_col):
                df_copy[col] = df_copy[col].astype(str).str.strip()
            df_copy[valor_col] = pd.to_numeric(df_copy[valor_col], errors='coerce').fillna(0.0)
            return df_copy.set_index([linha_col, grupo_col, subgrupo_col, tipo_col])[valor_col]

        def _flatten_fc_details(detalhes_fc_item):
            mapping = {
                'faturamento_linha': 'fat_linha',
                'conversao_linha': 'conv_linha',
                'faturamento_individual': 'fat_ind',
                'conversao_individual': 'conv_ind',
                'rentabilidade': 'rentab',
                'retencao_clientes': 'retencao',
                'meta_fornecedor_1': 'forn1',
                'meta_fornecedor_2': 'forn2'
            }
            flattened = {}
            for comp, short in mapping.items():
                detalhes = detalhes_fc_item.get(comp, {}) if isinstance(detalhes_fc_item, dict) else {}
                real_val = detalhes.get('realizado')
                if comp == 'rentabilidade' and real_val is not None:
                    try:
                        rv = float(real_val)
                        if rv > 1 and rv <= 100:
                            real_val = rv / 100.0
                        else:
                            real_val = rv
                    except Exception:
                        pass
                flattened[f'peso_{short}'] = detalhes.get('peso')
                flattened[f'realizado_{short}'] = real_val
                flattened[f'meta_{short}'] = detalhes.get('meta')
                flattened[f'ating_{short}'] = detalhes.get('atingimento')
                flattened[f'ating_cap_{short}'] = detalhes.get('atingimento_cap')
                flattened[f'comp_fc_{short}'] = detalhes.get('componente_fc')
                if comp.startswith('meta_fornecedor'):
                    flattened[f'moeda_{short}'] = detalhes.get('moeda')
            return flattened

        df_analise = self.data.get('ANALISE_COMERCIAL_COMPLETA', pd.DataFrame())
        if df_analise is None or df_analise.empty:
            _debug(f"    [Reconc-Info] Processo {proc_id} não encontrado no Analise_Comercial_Completa.")
            return [], 0.0

        proc_str = str(proc_id).strip()
        proc_col = _match_column(df_analise, ['processo'])
        if proc_col is None:
            _debug(f"    [Reconc-Erro] Coluna 'processo' não encontrada para localizar {proc_id}.")
            return [], 0.0
        df_itens_processo = df_analise[df_analise[proc_col].astype(str).str.strip() == proc_str].copy()
        if df_itens_processo.empty:
            try:
                proc_int_str = str(int(float(proc_str)))
                df_itens_processo = df_analise[df_analise[proc_col].astype(str).str.strip() == proc_int_str].copy()
            except Exception:
                df_itens_processo = pd.DataFrame()
        if df_itens_processo.empty:
            _debug(f"    [Reconc-Info] Processo {proc_id} não encontrado no Analise_Comercial_Completa.")
            return [], 0.0

        data_col = _match_column(df_itens_processo, ['dt emissão', 'dt emissao', 'data emissão', 'data emissao'])
        if data_col is None:
            _debug(f"    [Reconc-Erro] Não foi possível identificar a coluna de data de emissão para {proc_id}.")
            return [], 0.0
        data_emissao = pd.to_datetime(df_itens_processo[data_col].iloc[0], dayfirst=True, errors='coerce')
        if pd.isna(data_emissao):
            _debug(f"    [Reconc-Erro] Não foi possível ler Dt Emissão para {proc_id}.")
            return [], 0.0
        mes_fat = int(data_emissao.month)
        ano_fat = int(data_emissao.year)
        _debug(f"    [Reconc-Info] Processando {proc_id} para Mês/Ano: {mes_fat:02d}/{ano_fat}")

        _debug(f"      Carregando dados realizados para {mes_fat:02d}/{ano_fat}...")
        try:
            try:
                df_fat_hist, df_conv_hist, df_fat_ytd_hist, df_ret_hist = preparar_dados_mensais.prepare_dataframes_for_month(mes_fat, ano_fat, data_path=self.base_path)
            except TypeError:
                df_fat_hist, df_conv_hist, df_fat_ytd_hist, df_ret_hist = preparar_dados_mensais.prepare_dataframes_for_month(mes_fat, ano_fat)
        except Exception as e:
            _debug(f"    [Reconc-Erro] Falha ao carregar dados históricos via preparar_dados_mensais para {mes_fat:02d}/{ano_fat}: {e}")
            return [], 0.0

        df_fat_hist = df_fat_hist if isinstance(df_fat_hist, pd.DataFrame) else pd.DataFrame()
        df_conv_hist = df_conv_hist if isinstance(df_conv_hist, pd.DataFrame) else pd.DataFrame()
        df_fat_ytd_hist = df_fat_ytd_hist if isinstance(df_fat_ytd_hist, pd.DataFrame) else pd.DataFrame()
        df_ret_hist = df_ret_hist if isinstance(df_ret_hist, pd.DataFrame) else pd.DataFrame()

        rent_dir = os.path.join(self.base_path, 'rentabilidades')
        rent_xlsx = f"rentabilidade_{mes_fat:02d}_{ano_fat}_agrupada.xlsx"
        rent_csv = f"rentabilidade_{mes_fat:02d}_{ano_fat}_agrupada.csv"
        rent_path_xlsx = os.path.join(rent_dir, rent_xlsx)
        rent_path_csv = os.path.join(rent_dir, rent_csv)
        rentab_filename_used = rent_xlsx
        try:
            if os.path.exists(rent_path_xlsx):
                df_rentab_hist = pd.read_excel(rent_path_xlsx, engine='openpyxl')
                rentab_filename_used = rent_xlsx
            elif os.path.exists(rent_path_csv):
                df_rentab_hist = pd.read_csv(rent_path_csv, sep=';')
                rentab_filename_used = rent_csv
            else:
                raise FileNotFoundError(f"Arquivo de rentabilidade não encontrado: {rent_xlsx}")
            _debug(f"      Sucesso ao carregar {rentab_filename_used}")
        except Exception as e:
            _debug(f"    [Reconc-Erro] Falha ao carregar rentabilidade histórica '{rentab_filename_used}': {e}. Rentabilidade será 0.")
            df_rentab_hist = pd.DataFrame()

        col_valor_fat = _match_column(df_fat_hist, ['valor realizado', 'valor_realizado', 'valor nf', 'faturamento'])
        col_linha_fat = _match_column(df_fat_hist, ['negocio', 'negócio', 'linha'])
        col_consultor_fat = _match_column(df_fat_hist, ['consultor interno', 'consultor'])
        col_valor_conv = _match_column(df_conv_hist, ['valor orçado', 'valor orcado', 'valor_orcado'])
        col_linha_conv = _match_column(df_conv_hist, ['negocio', 'negócio', 'linha'])
        col_consultor_conv = _match_column(df_conv_hist, ['consultor interno', 'consultor'])

        realizados_hist = {
            'faturamento_linha': _build_group_series(df_fat_hist, col_linha_fat, col_valor_fat),
            'faturamento_individual': _build_group_series(df_fat_hist, col_consultor_fat, col_valor_fat),
            'conversao_linha': _build_group_series(df_conv_hist, col_linha_conv, col_valor_conv),
            'conversao_individual': _build_group_series(df_conv_hist, col_consultor_conv, col_valor_conv),
            'rentabilidade': _build_rentabilidade_series(df_rentab_hist)
        }
        for key in ('faturamento_linha', 'faturamento_individual', 'conversao_linha', 'conversao_individual', 'rentabilidade'):
            if key not in realizados_hist or realizados_hist[key] is None:
                realizados_hist[key] = pd.Series(dtype=float)

        linha_col = _match_column(df_itens_processo, ['negocio', 'negócio', 'linha'])
        grupo_col = _match_column(df_itens_processo, ['grupo'])
        subgrupo_col = _match_column(df_itens_processo, ['subgrupo'])
        tipo_col = _match_column(df_itens_processo, ['tipo de mercadoria', 'tipo mercadoria'])
        cliente_col = _match_column(df_itens_processo, ['cliente'])
        consultor_col = _match_column(df_itens_processo, ['consultor interno', 'consultor'])
        repres_col = _match_column(df_itens_processo, ['representante-pedido', 'representante'])
        valor_item_col = _match_column(df_itens_processo, ['valor realizado', 'faturamento', 'valor nf'])
        codigo_col = _match_column(df_itens_processo, ['código produto', 'codigo produto', 'cod produto', 'cod_produto'])
        descricao_col = _match_column(df_itens_processo, ['descrição produto', 'descricao produto', 'descricao_produto'])

        df_itens_processo['Processo'] = proc_str
        df_itens_processo['Dt Emissão'] = pd.to_datetime(df_itens_processo[data_col], dayfirst=True, errors='coerce') if data_col else pd.NaT
        if linha_col and linha_col != 'Negócio':
            df_itens_processo['Negócio'] = df_itens_processo[linha_col]
        elif 'Negócio' not in df_itens_processo.columns:
            df_itens_processo['Negócio'] = ''
        if grupo_col and grupo_col != 'Grupo':
            df_itens_processo['Grupo'] = df_itens_processo[grupo_col]
        elif 'Grupo' not in df_itens_processo.columns:
            df_itens_processo['Grupo'] = ''
        if subgrupo_col and subgrupo_col != 'Subgrupo':
            df_itens_processo['Subgrupo'] = df_itens_processo[subgrupo_col]
        elif 'Subgrupo' not in df_itens_processo.columns:
            df_itens_processo['Subgrupo'] = ''
        if tipo_col and tipo_col != 'Tipo de Mercadoria':
            df_itens_processo['Tipo de Mercadoria'] = df_itens_processo[tipo_col]
        elif 'Tipo de Mercadoria' not in df_itens_processo.columns:
            df_itens_processo['Tipo de Mercadoria'] = ''
        if cliente_col and cliente_col != 'Cliente':
            df_itens_processo['Cliente'] = df_itens_processo[cliente_col]
        elif 'Cliente' not in df_itens_processo.columns:
            df_itens_processo['Cliente'] = ''
        if consultor_col and consultor_col != 'Consultor Interno':
            df_itens_processo['Consultor Interno'] = df_itens_processo[consultor_col]
        elif 'Consultor Interno' not in df_itens_processo.columns:
            df_itens_processo['Consultor Interno'] = ''
        if repres_col and repres_col != 'Representante-pedido':
            df_itens_processo['Representante-pedido'] = df_itens_processo[repres_col]
        elif 'Representante-pedido' not in df_itens_processo.columns:
            df_itens_processo['Representante-pedido'] = ''
        if valor_item_col is None:
            df_itens_processo['Valor Realizado'] = 0.0
        else:
            df_itens_processo['Valor Realizado'] = pd.to_numeric(df_itens_processo[valor_item_col], errors='coerce').fillna(0.0)
        if codigo_col and codigo_col != 'Cód Produto':
            df_itens_processo['Cód Produto'] = df_itens_processo[codigo_col]
        if descricao_col and descricao_col != 'Descrição Produto':
            df_itens_processo['Descrição Produto'] = df_itens_processo[descricao_col]

        df_rc = self.data.get('RETENCAO_CLIENTES', None)
        df_ytd = self.data.get('FATURADOS_YTD', None)
        _prev_realizado = getattr(self, 'realizado', None)
        # Aplicar os realizados históricos temporariamente para que o cálculo do FC
        # utilize os valores do mês do faturamento do processo.
        self.realizado = realizados_hist
        # Log mínimo para depurar por que nenhuma linha é gerada durante reconciliação
        try:
            if getattr(self, '_logger', None):
                self._logger.info(f"[Reconc-Debug] realizados_hist sizes: " + ", ".join(f"{k}={getattr(v, 'shape', 'series') or len(v)}" for k,v in realizados_hist.items()))
            else:
                _debug("[Reconc-Debug] realizados_hist keys: " + str(list(realizados_hist.keys())))
        except Exception:
            pass
        try:
            self.data['RETENCAO_CLIENTES'] = df_ret_hist.copy() if not df_ret_hist.empty else pd.DataFrame()
            if not df_fat_ytd_hist.empty:
                dt_ytd_col = _match_column(df_fat_ytd_hist, ['dt emissão', 'dt emissao', 'data emissão', 'data emissao'])
                if dt_ytd_col:
                    df_fat_ytd_hist[dt_ytd_col] = pd.to_datetime(df_fat_ytd_hist[dt_ytd_col], dayfirst=True, errors='coerce')
                    if dt_ytd_col != 'Dt Emissão':
                        df_fat_ytd_hist['Dt Emissão'] = df_fat_ytd_hist[dt_ytd_col]
            self.data['FATURADOS_YTD'] = df_fat_ytd_hist.copy() if not df_fat_ytd_hist.empty else pd.DataFrame()

            df_atribuicoes = self.data.get('ATRIBUICOES', pd.DataFrame()).copy()
            df_colabs = self.data.get('COLABORADORES', pd.DataFrame()).copy()
            if not df_atribuicoes.empty:
                for col in ('linha', 'grupo', 'subgrupo', 'tipo_mercadoria', 'colaborador', 'cargo'):
                    if col in df_atribuicoes.columns:
                        df_atribuicoes[f'__norm_{col}'] = df_atribuicoes[col].apply(_normalize_text)
                for col in ('__norm_linha', '__norm_grupo', '__norm_subgrupo', '__norm_tipo_mercadoria', '__norm_colaborador', '__norm_cargo'):
                    if col not in df_atribuicoes.columns:
                        df_atribuicoes[col] = ''
            if not df_colabs.empty:
                if 'nome_colaborador' in df_colabs.columns:
                    df_colabs['__norm_nome_colaborador'] = df_colabs['nome_colaborador'].apply(_normalize_text)
                else:
                    df_colabs['__norm_nome_colaborador'] = ''
                if 'cargo' in df_colabs.columns:
                    df_colabs['__norm_cargo'] = df_colabs['cargo'].apply(_normalize_text)
                else:
                    df_colabs['__norm_cargo'] = ''
                if 'tipo_cargo' in df_colabs.columns:
                    df_colabs['__norm_tipo_cargo'] = df_colabs['tipo_cargo'].apply(_normalize_text)
                else:
                    df_colabs['__norm_tipo_cargo'] = ''

            cargos_gestao_norm = set()
            if not df_colabs.empty:
                try:
                    tipo_gestao_norm = _normalize_text('Gestão')
                except Exception:
                    tipo_gestao_norm = 'GESTAO'
                cargos_gestao_norm = set(df_colabs[df_colabs['__norm_tipo_cargo'] == tipo_gestao_norm]['__norm_cargo'].tolist())
            if cargos_gestao_norm:
                df_atribuicoes_gestao = df_atribuicoes[df_atribuicoes['__norm_cargo'].isin(cargos_gestao_norm)].copy()
            else:
                df_atribuicoes_gestao = pd.DataFrame(columns=df_atribuicoes.columns)

            colab_info_map = {}
            if not df_colabs.empty and '__norm_nome_colaborador' in df_colabs.columns:
                for _, row_info in df_colabs.iterrows():
                    norm_name = row_info.get('__norm_nome_colaborador')
                    if norm_name and norm_name not in colab_info_map:
                        colab_info_map[norm_name] = row_info

            recebe_set_norm = {_normalize_text(nome) for nome in getattr(self, 'recebe_por_recebimento', set())}
            try:
                if getattr(self, '_logger', None):
                    self._logger.info(f"[Reconc-Debug] recebe_por_recebimento (norm): {sorted(list(recebe_set_norm))}")
                    self._logger.info(f"[Reconc-Debug] df_atribuicoes_gestao shape: {getattr(df_atribuicoes_gestao, 'shape', None)} df_colabs shape: {getattr(df_colabs, 'shape', None)}")
                    # checar presença de um nome esperado
                    for nome_t in list(getattr(self, 'recebe_por_recebimento', set()))[:5]:
                        norm = _normalize_text(nome_t)
                        present = not df_colabs[df_colabs.get('__norm_nome_colaborador','') == norm].empty if '__norm_nome_colaborador' in df_colabs.columns else False
                        self._logger.info(f"[Reconc-Debug] colaborador '{nome_t}' present in COLABORADORES? {present}")
                else:
                    _debug("[Reconc-Debug] recebe_por_recebimento (norm): " + str(sorted(list(recebe_set_norm))))
            except Exception:
                pass
            linhas_reconciliacao_detalhada = []

            for _, item in df_itens_processo.iterrows():
                contexto_item = {
                    'linha': item.get('Negócio'),
                    'grupo': item.get('Grupo'),
                    'subgrupo': item.get('Subgrupo'),
                    'tipo_mercadoria': item.get('Tipo de Mercadoria')
                }

                linha_norm = _normalize_text(contexto_item['linha'])
                grupo_norm = _normalize_text(contexto_item['grupo'])
                subgrupo_norm = _normalize_text(contexto_item['subgrupo'])
                tipo_norm = _normalize_text(contexto_item['tipo_mercadoria'])

                if not df_atribuicoes_gestao.empty:
                    atribuidos_gestao = df_atribuicoes_gestao[
                        (df_atribuicoes_gestao['__norm_linha'] == linha_norm) &
                        (df_atribuicoes_gestao['__norm_grupo'] == grupo_norm) &
                        (df_atribuicoes_gestao['__norm_subgrupo'] == subgrupo_norm) &
                        (df_atribuicoes_gestao['__norm_tipo_mercadoria'] == tipo_norm)
                    ][['colaborador', 'cargo']].copy()
                else:
                    atribuidos_gestao = pd.DataFrame(columns=['colaborador', 'cargo'])

                nomes_operacionais = []
                for col_name in ('Consultor Interno', 'Representante-pedido'):
                    valor = item.get(col_name)
                    if valor is not None and str(valor).strip():
                        nomes_operacionais.append(str(valor).strip())
                nomes_oper_norm = {_normalize_text(n) for n in nomes_operacionais}
                atribuidos_operacional = df_colabs[df_colabs['__norm_nome_colaborador'].isin(nomes_oper_norm)] if (nomes_oper_norm and not df_colabs.empty and '__norm_nome_colaborador' in df_colabs.columns) else pd.DataFrame(columns=['nome_colaborador', 'cargo'])

                gestao_df = atribuidos_gestao[['colaborador', 'cargo']].copy() if not atribuidos_gestao.empty else pd.DataFrame(columns=['colaborador', 'cargo'])
                if not gestao_df.empty:
                    gestao_df['gestor'] = True
                operacional_df = atribuidos_operacional[['nome_colaborador', 'cargo']].rename(columns={'nome_colaborador': 'colaborador'}).copy() if not atribuidos_operacional.empty else pd.DataFrame(columns=['colaborador', 'cargo'])
                if not operacional_df.empty:
                    operacional_df['gestor'] = False

                combined = pd.concat([gestao_df, operacional_df], ignore_index=True, sort=False)
                try:
                    if getattr(self, '_logger', None):
                        self._logger.info(f"[Reconc-Debug] combined candidates for item (proc={proc_str}): {combined.to_dict(orient='records')}")
                    else:
                        _debug(f"[Reconc-Debug] combined candidates for item (proc={proc_str}): {combined.to_dict(orient='records')}")
                except Exception:
                    pass
                if combined.empty:
                    continue
                combined['colaborador'] = combined['colaborador'].astype(str).str.strip()
                combined['cargo'] = combined['cargo'].astype(str).str.strip()
                combined['__norm_colaborador'] = combined['colaborador'].apply(_normalize_text)
                combined = combined.drop_duplicates(subset=['__norm_colaborador', 'cargo']).reset_index(drop=True)
                combined = combined[combined['__norm_colaborador'].isin(recebe_set_norm)]
                # Se, após filtrar pelos que recebem por recebimento, não houver candidatos,
                # tentar trazer candidatos diretamente de df_colabs (fallback mínimo).
                # Para reconciliações, NUNCA usar fallback amplo de COLABORADORES.
                # Se não houver atribuídos (gestão/operacional) que também recebam por recebimento,
                # então não há cálculo de reconciliação para este item.
                if combined.empty:
                    if getattr(self, '_logger', None):
                        try:
                            self._logger.info(f"[Reconc-Debug] Sem candidatos elegíveis por recebimento para item (proc={proc_str}). Pulando.")
                        except Exception:
                            pass
                    continue

                processed = set()
                for _, colab_row in combined.iterrows():
                    norm_name = colab_row['__norm_colaborador']
                    if norm_name in processed:
                        continue
                    processed.add(norm_name)

                    colab_nome = colab_row['colaborador']
                    colab_cargo = colab_row['cargo']
                    regra = self._get_regra_comissao(**contexto_item, cargo=colab_cargo)
                    if regra is None:
                        continue
                    pe = _safe_float(regra.get('fatia_cargo_pct', 0.0)) / 100.0
                    taxa_rateio = _safe_float(regra.get('taxa_rateio_maximo_pct', 0.0)) / 100.0
                    faturamento_item = _safe_float(item.get('Valor Realizado', 0.0))
                    comissao_base = faturamento_item * taxa_rateio * pe

                    try:
                        fator_correcao_final, detalhes_fc_item = self._calcular_fc_para_item(colab_nome, colab_cargo, item)
                    except Exception as e:
                        _info(f"    [Reconc-Erro] Falha ao calcular FC para {colab_nome} no processo {proc_id}: {e}")
                        continue

                    comissao_calculada = comissao_base * fator_correcao_final
                    colab_info = colab_info_map.get(norm_name, {})
                    id_col = colab_info.get('id_colaborador') if isinstance(colab_info, pd.Series) else None
                    tipo_cargo_info = colab_info.get('tipo_cargo') if isinstance(colab_info, pd.Series) else None
                    if not isinstance(tipo_cargo_info, str) or not tipo_cargo_info.strip():
                        tipo_cargo_info = 'Gestão' if bool(colab_row.get('gestor')) else 'Operacional'
                    cargo_final = colab_info.get('cargo') if isinstance(colab_info, pd.Series) and isinstance(colab_info.get('cargo'), str) and colab_info.get('cargo').strip() else colab_cargo

                    linha_detalhada = {
                        'processo': proc_str,
                        'Dt Emissão': item.get('Dt Emissão'),
                        'Cliente': item.get('Cliente'),
                        'Negócio': item.get('Negócio'),
                        'Grupo': item.get('Grupo'),
                        'Subgrupo': item.get('Subgrupo'),
                        'Tipo de Mercadoria': item.get('Tipo de Mercadoria'),
                        'faturamento_item': faturamento_item,
                        'Faturamento': faturamento_item,
                        'comissao_potencial_maxima': comissao_base,
                        'comissao_base': comissao_base,
                        'comissao_calculada': comissao_calculada,
                        'fator_correcao_fc': fator_correcao_final,
                        'fator_correcao_final': fator_correcao_final,
                        'percentual_elegibilidade_pe': pe,
                        'pe_aplicado': pe,
                        'taxa_rateio_aplicada': taxa_rateio,
                        'id_colaborador': id_col,
                        'colaborador': colab_nome,
                        'cargo': cargo_final,
                        'tipo_cargo': tipo_cargo_info,
                        'cod_produto': item.get('Cód Produto'),
                        'descricao_produto': item.get('Descrição Produto')
                    }
                    linha_detalhada.update(_flatten_fc_details(detalhes_fc_item))
                    linhas_reconciliacao_detalhada.append(linha_detalhada)
        finally:
            # Restaurar o atributo realizado e outros dados originais
            self.realizado = _prev_realizado
            self.data['RETENCAO_CLIENTES'] = df_rc
            self.data['FATURADOS_YTD'] = df_ytd

        comissao_correta_total = sum(_safe_float(row.get('comissao_calculada', 0.0)) for row in linhas_reconciliacao_detalhada)
        return linhas_reconciliacao_detalhada, comissao_correta_total

    def _get_regra_comissao(self, linha, grupo, subgrupo, tipo_mercadoria, cargo):
        """Busca a regra de comissão aplicável considerando hierarquia de especificidade."""
        chave_cache = (linha, grupo, subgrupo, tipo_mercadoria, cargo)
        if chave_cache in self.cache_regras:
            return self.cache_regras[chave_cache]

        df_regras = self.data.get('CONFIG_COMISSAO', pd.DataFrame())
        if df_regras.empty:
            self._log_validacao('ERRO', "Tabela CONFIG_COMISSAO indisponível para cálculo de regras.", {})
            self.cache_regras[chave_cache] = None
            return None

        filtros = [
            (df_regras['linha'] == linha) & (df_regras['grupo'] == grupo) & (df_regras['subgrupo'] == subgrupo) & (df_regras['tipo_mercadoria'] == tipo_mercadoria),
            (df_regras['linha'] == linha) & (df_regras['grupo'] == grupo) & ((df_regras['subgrupo'].isna()) | (df_regras['subgrupo'] == self.legacy_token)) & (df_regras['tipo_mercadoria'] == tipo_mercadoria),
            (df_regras['linha'] == linha) & ((df_regras['grupo'].isna()) | (df_regras['grupo'] == self.legacy_token)) &
            ((df_regras['subgrupo'].isna()) | (df_regras['subgrupo'] == self.legacy_token)) & (df_regras['tipo_mercadoria'] == tipo_mercadoria),
            (df_regras['linha'] == self.legacy_token) & (df_regras['tipo_mercadoria'] == self.legacy_token)
        ]

        for filtro in filtros:
            try:
                regra = df_regras[filtro & (df_regras['cargo'] == cargo)]
            except Exception:
                regra = pd.DataFrame()
            if not regra.empty:
                regra_row = regra.iloc[0]
                self.cache_regras[chave_cache] = regra_row
                return regra_row

        self._log_validacao('ERRO', 'Nenhuma regra de comissão encontrada.', {
            "linha": linha, "grupo": grupo, "subgrupo": subgrupo, "tipo_mercadoria": tipo_mercadoria, "cargo": cargo
        })
        self.cache_regras[chave_cache] = None
        return None

    def _calcular_comissoes(self):
        """Itera sobre os itens faturados, calcula o FC para cada um e a comissão final."""
        comissoes_calculadas = []
        # auditoria detalhada agora é armazenada nas colunas de COMISSOES_CALCULADAS
        df_faturados = self.data['FATURADOS']
        df_atribuicoes = self.data['ATRIBUICOES']
        df_colabs_com_cargos = self.data['COLABORADORES']

        # Pre-filtra atribuições de gestão para otimização
        cargos_gestao = df_colabs_com_cargos[df_colabs_com_cargos['tipo_cargo'] == 'Gestão']['cargo'].unique()
        df_atribuicoes_gestao = df_atribuicoes[df_atribuicoes['cargo'].isin(cargos_gestao)]

        # --- Detecção de Cross-Selling por Processo (pré-scan) ---
        # Estrutura: self.cross_selling_decisions[processo] = {is_cross:bool, consultor:str, linha:str, taxa:float, decision:'A'|'B'}
        self.cross_selling_decisions = {}
        try:
            # Construir mapa de aliases para colaboradores (case-insensitive)
            alias_map = {}
            alias_map_lower = {}
            if 'ALIASES' in self.data and not self.data['ALIASES'].empty:
                aliases_df = self.data['ALIASES'][self.data['ALIASES']['entidade'] == 'colaborador'][['alias','padrao']].dropna()
                for _, r in aliases_df.iterrows():
                    a = str(r['alias']).strip()
                    p = str(r['padrao']).strip()
                    alias_map[a] = p
                    alias_map_lower[a.lower()] = p

            cross_df = self.data.get('CROSS_SELLING', pd.DataFrame())

            if 'Processo' in df_faturados.columns:
                for processo, grupo in df_faturados.groupby('Processo'):
                    primeira = grupo.iloc[0]
                    gerente_comercial_raw = None
                    if 'Gerente Comercial-Pedido' in primeira.index:
                        gerente_comercial_raw = primeira.get('Gerente Comercial-Pedido')
                    if gerente_comercial_raw is None or pd.isna(gerente_comercial_raw) or str(gerente_comercial_raw).strip() == '':
                        continue

                    raw_norm = str(gerente_comercial_raw).strip()
                    gerente_padrao = alias_map.get(raw_norm)
                    if gerente_padrao is None:
                        gerente_padrao = alias_map_lower.get(raw_norm.lower(), raw_norm)

                    # Verificar se é Consultor Externo
                    # matching case-insensitive against coluna 'nome_colaborador'
                    try:
                        mask_col = df_colabs_com_cargos['nome_colaborador'].astype(str).str.strip().str.lower() == str(gerente_padrao).strip().lower()
                        row_colab = df_colabs_com_cargos[mask_col]
                    except Exception:
                        row_colab = df_colabs_com_cargos[df_colabs_com_cargos['nome_colaborador'] == gerente_padrao]
                    if row_colab.empty:
                        continue
                    cargo_do_consultor = row_colab.iloc[0].get('cargo', '')
                    tipo_cargo = row_colab.iloc[0].get('tipo_cargo', '')
                    is_consultor_externo = (str(cargo_do_consultor).strip().lower() == 'consultor externo') or (str(tipo_cargo).strip().lower() == 'externo')
                    if not is_consultor_externo:
                        continue

                    # Determinar a linha do processo (usar a primeira linha encontrada)
                    linhas_no_processo = grupo['Negócio'].dropna().unique().tolist()
                    if not linhas_no_processo:
                        continue
                    linha_do_processo = linhas_no_processo[0]

                    # Verificar se o consultor possui atribuições para esta linha
                    possui_atr = False
                    if not df_atribuicoes.empty:
                        possui_atr = not df_atribuicoes[(df_atribuicoes['colaborador'] == gerente_padrao) & (df_atribuicoes['linha'] == linha_do_processo)].empty

                    if not possui_atr:
                        # Consultor externo não possui atribuição para esta linha -> cross-selling detectado
                        taxa = 0.0
                        try:
                            if not cross_df.empty:
                                # case-insensitive match
                                mask_cs = cross_df['colaborador'].astype(str).str.strip().str.lower() == str(gerente_padrao).strip().lower()
                                row_cs = cross_df[mask_cs]
                                if not row_cs.empty:
                                    taxa = float(row_cs.iloc[0].get('taxa_cross_selling_pct', 0.0))
                        except Exception:
                            taxa = 0.0

                        # Obter decisão do usuário (prompt) — usar opção default quando não interativo
                        try:
                            decisao = self._handle_cross_selling_prompt(processo, gerente_padrao, linha_do_processo, taxa)
                        except Exception:
                            decisao = self.params.get('cross_selling_default_option', 'A')

                        self.cross_selling_decisions[processo] = {
                            'is_cross': True,
                            'consultor': gerente_padrao,
                            'linha': linha_do_processo,
                            'taxa': float(taxa),
                            'decision': decisao,
                            'timestamp': datetime.now().isoformat()
                        }
        except Exception as e:
            self._log_validacao('AVISO', f'Erro na detecção de cross-selling: {e}', {})

        try:
            total_items_step5 = len(df_faturados) if df_faturados is not None else 0
        except Exception:
            total_items_step5 = 0
        processed_step5 = 0
        progress_step5_mod = max(1, total_items_step5 // 100) if total_items_step5 else 1
        _progress_step5(0, total_items_step5)

        for _, item_faturado in df_faturados.iterrows():
            processed_step5 += 1
            if (
                processed_step5 == total_items_step5
                or processed_step5 % progress_step5_mod == 0
                or processed_step5 % 100 == 0
            ):
                _progress_step5(processed_step5, total_items_step5)
            contexto_item = {
                'linha': item_faturado['Negócio'], 'grupo': item_faturado['Grupo'],
                'subgrupo': item_faturado['Subgrupo'], 'tipo_mercadoria': item_faturado['Tipo de Mercadoria']
            }
            
            # 1. Obter time de GESTÃO a partir das ATRIBUICOES
            atribuidos_gestao = df_atribuicoes_gestao[
                (df_atribuicoes_gestao['linha'] == contexto_item['linha']) &
                (df_atribuicoes_gestao['grupo'] == contexto_item['grupo']) &
                (df_atribuicoes_gestao['subgrupo'] == contexto_item['subgrupo']) &
                (df_atribuicoes_gestao['tipo_mercadoria'] == contexto_item['tipo_mercadoria'])
            ]

            # 2. Obter time OPERACIONAL a partir do item FATURADO
            nomes_operacionais = []
            if pd.notna(item_faturado.get('Consultor Interno')):
                nomes_operacionais.append(item_faturado['Consultor Interno'])
            if pd.notna(item_faturado.get('Representante-pedido')):
                nomes_operacionais.append(item_faturado['Representante-pedido'])
            
            atribuidos_operacional = df_colabs_com_cargos[
                df_colabs_com_cargos['nome_colaborador'].isin(nomes_operacionais)
            ]

            # 3. Combinar os times
            # Normalizar e combinar listas de colaboradores; garantir que nomes iguais e cargos iguais
            # resultem em apenas uma entrada. Em alguns casos, pequenas diferenças de whitespace/maiusculas
            # podem impedir que drop_duplicates remova as duplicatas, então normalizamos os nomes
            # e aplicamos deduplicação por 'colaborador' e 'cargo'. Além disso mantemos um conjunto
            # processed_colabs durante a iteração para garantir que cada colaborador seja processado
            # no máximo uma vez por item_faturado.
            gestion = atribuidos_gestao[['colaborador', 'cargo']].copy() if not atribuidos_gestao.empty else pd.DataFrame(columns=['colaborador','cargo'])
            operacional = atribuidos_operacional[['nome_colaborador', 'cargo']].rename(columns={'nome_colaborador': 'colaborador'}).copy() if not atribuidos_operacional.empty else pd.DataFrame(columns=['colaborador','cargo'])
            # Normalizar texto (strip e lower) para comparação e deduplicação
            for df_tmp in (gestion, operacional):
                if not df_tmp.empty and 'colaborador' in df_tmp.columns:
                    df_tmp['colaborador'] = df_tmp['colaborador'].astype(str).str.strip()
                    # preservar original-case in the final DataFrame, but dedupe on normalized
            combined = pd.concat([gestion, operacional], ignore_index=True, sort=False)
            # deduplicate by colaborador and cargo after normalization of whitespace
            combined['__colab_norm'] = combined['colaborador'].astype(str).str.lower().str.strip()
            combined = combined.drop_duplicates(subset=['__colab_norm', 'cargo']).drop(columns=['__colab_norm']).reset_index(drop=True)
            colaboradores_para_comissionar = combined

            # Fallback mínimo: se, após aplicar o filtro por quem recebe por recebimento,
            # não houver candidatos, tentar incluir colaboradores listados em
            # `recebe_por_recebimento` que existam em `df_colabs`.
            try:
                if colaboradores_para_comissionar.empty and recebe_set_norm and not df_colabs.empty:
                    candidatos = df_colabs[df_colabs['__norm_nome_colaborador'].isin(recebe_set_norm)][['nome_colaborador', 'cargo']].copy()
                    if not candidatos.empty:
                        candidatos = candidatos.rename(columns={'nome_colaborador': 'colaborador'})
                        candidatos['colaborador'] = candidatos['colaborador'].astype(str).str.strip()
                        candidatos['cargo'] = candidatos['cargo'].astype(str).str.strip()
                        candidatos['__norm_colaborador'] = candidatos['colaborador'].apply(_normalize_text)
                        candidatos = candidatos.drop_duplicates(subset=['__norm_colaborador', 'cargo']).drop(columns=['__norm_colaborador']).reset_index(drop=True)
                        colaboradores_para_comissionar = candidatos
                        if getattr(self, '_logger', None):
                            self._logger.info(f"[Reconc-Debug] Fallback: usando colaboradores de COLABORADORES para reconciliacao: {colaboradores_para_comissionar.to_dict(orient='records')}")
                        else:
                            _debug("[Reconc-Debug] Fallback: usando colaboradores de COLABORADORES para reconciliacao: " + str(colaboradores_para_comissionar.to_dict(orient='records')))
            except Exception:
                pass


            # Verificar se este processo foi detectado como cross-selling
            processo_atual = item_faturado.get('Processo') if 'Processo' in item_faturado.index else None
            cs_info = self.cross_selling_decisions.get(processo_atual, None)

            # Se for cross-selling, gerar comissão especial para o consultor externo
            if cs_info and cs_info.get('is_cross'):
                consultor_externo = cs_info.get('consultor')
                taxa_cs = float(cs_info.get('taxa', 0.0)) / 100.0
                # Para cada item, calcular comissão especial e adicioná-la como uma linha distinta
                try:
                    if taxa_cs and taxa_cs > 0:
                        comissao_cs = item_faturado['Valor Realizado'] * taxa_cs
                        # identificar id_colaborador se existir
                        id_col = None
                        row_col = df_colabs_com_cargos[df_colabs_com_cargos['nome_colaborador'] == consultor_externo]
                        if not row_col.empty:
                            id_col = row_col.iloc[0].get('id_colaborador')

                        comissoes_calculadas.append({
                            'id_colaborador': id_col,
                            'nome_colaborador': consultor_externo,
                            'cargo': 'Consultor Externo',
                            'cod_produto': item_faturado.get('Código Produto'),
                            'descricao_produto': item_faturado.get('Descrição Produto'),
                            'processo': processo_atual,
                            'linha': item_faturado.get('Negócio'),
                            'grupo': item_faturado.get('Grupo'),
                            'subgrupo': item_faturado.get('Subgrupo'),
                            'tipo_mercadoria': item_faturado.get('Tipo de Mercadoria'),
                            'faturamento_item': item_faturado.get('Valor Realizado'),
                            'taxa_rateio_aplicada': None,
                            'fator_correcao_fc': 1.0,
                            'percentual_elegibilidade_pe': None,
                            'comissao_potencial_maxima': None,
                            'comissao_calculada': comissao_cs,
                            'observacao': 'CROSS_SELLING'
                        })
                except Exception:
                    pass
            if colaboradores_para_comissionar.empty:
                self._log_validacao('AVISO', "Nenhum colaborador (gestão ou operacional) encontrado para o item.", dict(item_faturado))
                continue

            # runtime guard: ensure each collaborator is processed at most once per item
            processed_colabs = set()
            for _, atribuicao in colaboradores_para_comissionar.iterrows():
                colab_nome = atribuicao['colaborador']
                colab_cargo = atribuicao['cargo']
                # build normalized key to detect duplicates robustly
                key_colab = (str(colab_nome).strip().lower(), str(colab_cargo).strip().lower(),
                             str(processo_atual) if processo_atual is not None else '',
                             str(item_faturado.get('Código Produto', '')).strip().lower())
                if key_colab in processed_colabs:
                    # já processado para este item; pular
                    continue
                processed_colabs.add(key_colab)
                
                # Se este colaborador for o Consultor Externo removido na opção B, pular
                if cs_info and cs_info.get('is_cross') and cs_info.get('decision') == 'B':
                    # opção B: o Consultor Externo é removido do cálculo normal
                    if colab_nome == cs_info.get('consultor'):
                        continue

                regra = self._get_regra_comissao(**contexto_item, cargo=colab_cargo)
                if regra is None: continue
                fc, detalhes_fc_item = self._calcular_fc_para_item(colab_nome, colab_cargo, item_faturado)

                faturamento_item = item_faturado['Valor Realizado']
                taxa_rateio = regra['taxa_rateio_maximo_pct'] / 100.0
                pe = regra['fatia_cargo_pct'] / 100.0

                # Se este processo tem cross-selling e a decisão foi A (subtrair), reduzir taxa_rateio
                if cs_info and cs_info.get('is_cross') and cs_info.get('decision') == 'A':
                    # reduzir taxa em taxa_cs percentual (taxa armazenada em %)
                    taxa_reduc = float(cs_info.get('taxa', 0.0)) / 100.0
                    # Avisar se taxa_cross_selling_pct > taxa_rateio_maximo_pct
                    try:
                        if taxa_reduc > regra['taxa_rateio_maximo_pct'] / 100.0:
                            self._log_validacao('AVISO', f"taxa_cross_selling_pct ({taxa_reduc:.4f}) maior que taxa_rateio_maximo_pct ({regra['taxa_rateio_maximo_pct']/100.0:.4f}) para processo {processo_atual}", {'processo': processo_atual, 'consultor': cs_info.get('consultor'), 'taxa_cs': taxa_reduc, 'taxa_rateio': regra['taxa_rateio_maximo_pct']/100.0})
                    except Exception:
                        pass
                    taxa_rateio = max(0.0, taxa_rateio - taxa_reduc)

                # Se a decisão for B, mantemos taxa_rateio intacta mas consultor externo já foi removido
                
                comissao_potencial = faturamento_item * taxa_rateio * pe
                comissao_item = comissao_potencial * fc
                
                # construir dicionário base e depois anexar colunas detalhadas do FC
                base_dict = {
                    'id_colaborador': df_colabs_com_cargos.loc[df_colabs_com_cargos['nome_colaborador'] == colab_nome, 'id_colaborador'].iloc[0],
                    'nome_colaborador': colab_nome, 'cargo': colab_cargo,
                    'cod_produto': item_faturado['Código Produto'], 'descricao_produto': item_faturado['Descrição Produto'],
                    'processo': item_faturado['Processo'], **contexto_item,
                    'faturamento_item': faturamento_item, 'taxa_rateio_aplicada': taxa_rateio,
                    'fator_correcao_fc': fc, 'percentual_elegibilidade_pe': pe,
                    'comissao_potencial_maxima': comissao_potencial,
                    'comissao_calculada': comissao_item
                }

                # helper para extrair valores seguros do detalhes_fc_item
                def _g(dct, key, subkey, default=None):
                    try:
                        v = dct.get(key)
                        if v is None:
                            return default
                        return v.get(subkey, default)
                    except Exception:
                        return default

                # Mapear componentes padronizados para colunas
                mapping = {
                    'faturamento_linha': 'fat_linha',
                    'conversao_linha': 'conv_linha',
                    'faturamento_individual': 'fat_ind',
                    'conversao_individual': 'conv_ind',
                    'rentabilidade': 'rentab'
                }

                for comp, short in mapping.items():
                    detalhes = detalhes_fc_item.get(comp) if isinstance(detalhes_fc_item, dict) else None
                    base_dict[f'peso_{short}'] = _g(detalhes_fc_item, comp, 'peso', None)
                    # Normalizar rentabilidade: garantir que realizado (rentab) esteja em decimal (ex: 0.12)
                    real_val = _g(detalhes_fc_item, comp, 'realizado', None)
                    if comp == 'rentabilidade' and real_val is not None:
                        try:
                            # se valor aparenta estar em porcentagem (>1 e <=100), converter dividindo por 100
                            rv = float(real_val)
                            if rv > 1 and rv <= 100:
                                rv = rv / 100.0
                            real_val = rv
                        except Exception:
                            pass
                    base_dict[f'realizado_{short}'] = real_val
                    base_dict[f'meta_{short}'] = _g(detalhes_fc_item, comp, 'meta', None)
                    # Atingimento é uma razão (realizado/meta) e deve ser mantido como está (pode ser >1)
                    base_dict[f'ating_{short}'] = _g(detalhes_fc_item, comp, 'atingimento', None)
                    base_dict[f'ating_cap_{short}'] = _g(detalhes_fc_item, comp, 'atingimento_cap', None)
                    base_dict[f'comp_fc_{short}'] = _g(detalhes_fc_item, comp, 'componente_fc', None)
                    # se houver moeda (aplicável a fornecedores), incluir coluna moeda_
                    if comp.startswith('meta_fornecedor'):
                        base_dict[f'moeda_{short}'] = _g(detalhes_fc_item, comp, 'moeda', None)

                # Após popular todas as colunas de detalhe do FC, anexar a linha apenas uma vez
                comissoes_calculadas.append(base_dict)

                self.comissoes_df = pd.DataFrame(comissoes_calculadas)

        if total_items_step5 and processed_step5 < total_items_step5:
            _progress_step5(total_items_step5, total_items_step5)
        elif total_items_step5 == 0:
            sys.stdout.write("\n")
        _info(f"[Resumo] Etapa 5 concluída: {processed_step5} itens processados.")

    def _handle_cross_selling_prompt(self, processo, consultor, linha, taxa):
        """Mostra prompt interativo no terminal para decisão A ou B sobre o cross-selling.

        Retorna 'A' ou 'B'. Usa cross_selling_default_option quando a entrada for vazia (modo não interativo).
        """
        default = self.params.get('cross_selling_default_option', 'A')
        prompt = f"\n------------------------------------------------------------------\n"
        prompt += "[!] ALERTA DE CROSS-SELLING DETECTADO\n\n"
        prompt += f"Processo: {processo}\n"
        prompt += f"Consultor Externo: {consultor}\n"
        prompt += f"Linha da Venda: {linha} (fora de sua carteira)\n\n"
        prompt += f"Escolha como a comissão de Cross-Selling de {taxa}% será tratada:\n\n"
        prompt += "(A) SUBTRAIR da Taxa de Rateio:\n"
        prompt += "    - A comissão de Cross-Selling será paga ao Consultor Externo.\n"
        prompt += f"    - A Taxa de Rateio dos itens deste processo será REDUZIDA em {taxa}% para os demais colaboradores.\n\n"
        prompt += "(B) PAGAR SEPARADAMENTE:\n"
        prompt += "    - A comissão de Cross-Selling será paga ao Consultor Externo.\n"
        prompt += "    - A Taxa de Rateio dos itens permanecerá INTACTA para os demais.\n"
        prompt += "    - O Consultor Externo será REMOVIDO do cálculo de comissão normal para este processo.\n\n"
        prompt += "Digite sua escolha (A ou B) e pressione Enter: "

        # Se rodando em ambiente não-interativo, usar default
        try:
            # Se stdin não for interativo, retornar o default
            if not sys.stdin.isatty():
                if getattr(self, '_logger', None):
                    self._logger.info(f"Modo não-interativo detectado — usando opção default de cross-selling: {default}")
                return default
        except Exception:
            pass
        while True:
            try:
                escolha = input(prompt).strip().upper()
            except Exception:
                escolha = ''

            if escolha == '':
                escolha = default

            if escolha in ('A', 'B'):
                return escolha
            else:
                print("Entrada inválida. Digite 'A' ou 'B'.")
    
    def _gerar_detalhamento_pdf(self):
        """Gera um PDF detalhando o cálculo de cada comissão."""
        if not REPORTLAB_DISPONIVEL:
            return

        # Importa explicitamente as dependências do reportlab aqui para garantir que estejam disponíveis
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet

        df_comissoes = self.comissoes_df
        # Se não existirem itens faturados no mês selecionado, garantir que as comissões
        # de faturamento sejam zero e registrar um aviso na planilha de saída.
        try:
            df_faturados = self.data.get('FATURADOS', pd.DataFrame())
            # Detectar se há linhas com 'Dt Emissão' no mês/ano atual (quando coluna presente)
            tem_faturamento_no_mes = True
            if 'Dt Emissão' in df_faturados.columns and not df_faturados.empty:
                try:
                    # usar o parâmetro mes/ano passado via self.params se disponível
                    raw_mes = self.params.get('mes_apuracao') if isinstance(self.params, dict) else None
                    raw_ano = self.params.get('ano_apuracao') if isinstance(self.params, dict) else None
                    mes_param = int(raw_mes) if raw_mes not in (None, '', False) else None
                    ano_param = int(raw_ano) if raw_ano not in (None, '', False) else None
                except Exception:
                    mes_param = None
                    ano_param = None

                if mes_param and ano_param:
                    df_dates = pd.to_datetime(df_faturados['Dt Emissão'], errors='coerce')
                    df_faturados_mes = df_faturados[(df_dates.dt.month == mes_param) & (df_dates.dt.year == ano_param)]
                    tem_faturamento_no_mes = not df_faturados_mes.empty
                else:
                    tem_faturamento_no_mes = not df_faturados.empty
            elif df_faturados.empty:
                tem_faturamento_no_mes = False

            if not tem_faturamento_no_mes:
                # Zeroar comissões de faturamento
                if not df_comissoes.empty:
                    df_comissoes['comissao_calculada'] = 0.0
                # Adicionar aviso de falta de faturamento
                self._log_validacao('AVISO', 'Nenhum item faturado encontrado para o mês selecionado; comissões de faturamento zeradas.', {})
                # Inserir linha de aviso quando for gerar o Excel (mais abaixo) — marcar via atributo
                self._no_faturamento_mes = True
            else:
                self._no_faturamento_mes = False
        except Exception:
            # se algo falhar, não impedir a geração, apenas continuar
            self._no_faturamento_mes = False
        # Detalhes do FC agora estão incorporados nas colunas de df_comissoes

        # --- Limpeza local apenas para apresentação no PDF ---
        # Evita páginas duplicadas no PDF quando o DataFrame de comissões
        # contém linhas repetidas (isso não altera os dados usados nos
        # cálculos ou na geração do Excel)
        if not df_comissoes.empty:
            subset_cols = [
                'nome_colaborador', 'cod_produto', 'processo', 'descricao_produto', 'cargo',
                'linha', 'grupo', 'subgrupo', 'tipo_mercadoria',
                'faturamento_item', 'taxa_rateio_aplicada', 'percentual_elegibilidade_pe', 'fator_correcao_fc',
                'comissao_potencial_maxima', 'comissao_calculada'
            ]
            # remove duplicatas aparentes só para o relatório
            existing_cols = [c for c in subset_cols if c in df_comissoes.columns]
            df_comissoes = df_comissoes.drop_duplicates(subset=existing_cols)

        if df_comissoes.empty:
            return

        # Segurança: se houver muitas comissões, gerar o PDF apenas com uma amostra
        MAX_PAGES_PDF = int(self.params.get('max_pages_pdf', 200)) if isinstance(self.params.get('max_pages_pdf', None), (int, float, str)) else 200
        if len(df_comissoes) > MAX_PAGES_PDF:
            # Amostragem estratificada simples por (linha, cargo)
            try:
                sample_size = int(self.params.get('sample_pages_pdf', 100)) if self.params.get('sample_pages_pdf') else 100
            except Exception:
                sample_size = 100
            # Construir pequena amostra para relatório
            grp = df_comissoes.groupby(['linha', 'cargo'], dropna=False)
            samples = []
            for _, g in grp:
                try:
                    samples.append(g.sample(1, random_state=42))
                except Exception:
                    continue
            if samples:
                df_comissoes = pd.concat(samples).head(sample_size)
            else:
                df_comissoes = df_comissoes.sample(min(sample_size, len(df_comissoes)), random_state=42)

        nome_arquivo_pdf = f"Detalhamento_Comissoes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(nome_arquivo_pdf)
        styles = getSampleStyleSheet()
        story = []
        
        for index, comissao in df_comissoes.iterrows():
            story.append(Paragraph("<b>Detalhamento do Cálculo de Comissão</b>", styles['h1']))
            story.append(Spacer(1, 12))
            
            story.append(Paragraph(f"<b>Processo:</b> {comissao['processo']}", styles['Normal']))
            story.append(Paragraph(f"<b>Produto:</b> {comissao['cod_produto']} - {comissao['descricao_produto']}", styles['Normal']))
            story.append(Paragraph(f"<b>Colaborador:</b> {comissao['nome_colaborador']} ({comissao['cargo']})", styles['Normal']))
            story.append(Paragraph(f"<b>Faturamento do Item:</b> R$ {comissao['faturamento_item']:.2f}", styles['Normal']))
            story.append(Spacer(1, 24))
            
            story.append(Paragraph("<b>Passo 1: Aplicação da Regra de Comissão</b>", styles['h2']))
            story.append(Paragraph(f"Para o contexto (Linha: {comissao['linha']}, ..., Cargo: {comissao['cargo']}), a regra encontrada foi:", styles['Normal']))
            story.append(Paragraph(f"&nbsp;&nbsp;&nbsp;- Taxa de Rateio (taxa_rateio_aplicada): {comissao['taxa_rateio_aplicada']:.2%}", styles['Normal']))
            story.append(Paragraph(f"&nbsp;&nbsp;&nbsp;- Percentual de Elegibilidade (PE): {comissao['percentual_elegibilidade_pe']:.2%}", styles['Normal']))
            story.append(Spacer(1, 24))

            story.append(Paragraph("<b>Passo 2: Cálculo do Fator de Correção (FC)</b>", styles['h2']))
            
            # Detalhes do FC agora estão nas colunas do DataFrame de comissões. Imprimir por tipo se o peso existir.
            mapping = {
                'faturamento_linha': 'fat_linha',
                'conversao_linha': 'conv_linha',
                'faturamento_individual': 'fat_ind',
                'conversao_individual': 'conv_ind',
                'rentabilidade': 'rentab',
                'retencao_clientes': 'retencao',
                'meta_fornecedor_1': 'forn1',
                'meta_fornecedor_2': 'forn2'
            }

            any_comp = False
            for tipo_meta, short in mapping.items():
                peso_col = f'peso_{short}'
                if peso_col in comissao.index and comissao.get(peso_col) not in (None, 0, 0.0) and not pd.isna(comissao.get(peso_col)):
                    any_comp = True
                    peso = comissao.get(peso_col) or 0
                    realizado = comissao.get(f'realizado_{short}', None)
                    meta_val = comissao.get(f'meta_{short}', None)
                    atingimento = comissao.get(f'ating_{short}', 0) or 0
                    ating_cap = comissao.get(f'ating_cap_{short}', 0) or 0
                    componente_fc = comissao.get(f'comp_fc_{short}', 0) or 0

                    # formatar strings similares ao comportamento anterior
                    if pd.isna(meta_val):
                        meta_str = 'N/A'
                    else:
                        if 'fat' in short or 'conv' in short:
                            try:
                                meta_str = f"R$ {float(meta_val):,.2f}"
                            except Exception:
                                meta_str = str(meta_val)
                        else:
                            meta_str = f"{float(meta_val):.2f}" if isinstance(meta_val, (int, float, np.floating)) else str(meta_val)

                    if pd.isna(realizado):
                        realizado_str = 'N/A'
                    else:
                        if 'fat' in short or 'conv' in short:
                            try:
                                realizado_str = f"R$ {float(realizado):,.2f}"
                            except Exception:
                                realizado_str = str(realizado)
                        else:
                            realizado_str = f"{float(realizado):.2f}" if isinstance(realizado, (int, float, np.floating)) else str(realizado)

                    story.append(Paragraph(f"<b>Componente: {tipo_meta}</b>", styles['h3']))
                    story.append(Paragraph(f"&nbsp;&nbsp;&nbsp;- Peso da Meta: {peso:.2%}", styles['Normal']))
                    story.append(Paragraph(f"&nbsp;&nbsp;&nbsp;- Valor Realizado: {realizado_str}", styles['Normal']))
                    story.append(Paragraph(f"&nbsp;&nbsp;&nbsp;- Valor da Meta: {meta_str}", styles['Normal']))
                    story.append(Paragraph(f"&nbsp;&nbsp;&nbsp;- Atingimento: {atingimento:.2%}", styles['Normal']))
                    story.append(Paragraph(f"&nbsp;&nbsp;&nbsp;- Atingimento (com cap): {ating_cap:.2%}", styles['Normal']))
                    story.append(Paragraph(f"&nbsp;&nbsp;&nbsp;- <b>Cálculo do Componente FC:</b> {ating_cap:.2%} * {peso:.2%} = <b>{componente_fc:.4f}</b>", styles['Normal']))
                    story.append(Spacer(1, 6))

            if not any_comp:
                story.append(Paragraph("Nenhum componente de FC aplicável para este cargo/item.", styles['Normal']))

            story.append(Paragraph(f"<b>FC Total para este Item:</b> {comissao['fator_correcao_fc']:.4f}", styles['h3']))
            story.append(Spacer(1, 24))

            story.append(Paragraph("<b>Passo 3: Cálculo Final da Comissão</b>", styles['h2']))
            formula = "Faturamento * Taxa de Rateio * PE * FC"
            calculo = f"R$ {comissao['faturamento_item']:.2f} * {comissao['taxa_rateio_aplicada']:.2%} * {comissao['percentual_elegibilidade_pe']:.2%} * {comissao['fator_correcao_fc']:.4f}"
            story.append(Paragraph(f"<b>Potencial Máximo (FC=1.0): R$ {comissao['comissao_potencial_maxima']:.2f}</b>", styles['Normal']))
            story.append(Paragraph(f"<b>Fórmula Efetiva:</b> {formula}", styles['Normal']))
            story.append(Paragraph(f"<b>Cálculo Efetivo:</b> {calculo}", styles['Normal']))
            story.append(Paragraph(f"<b>Comissão Final (Efetiva) para este Item: R$ {comissao['comissao_calculada']:.2f}</b>", styles['h3']))
            
            story.append(PageBreak())
        
        try:
            doc.build(story)
            _info(f"PDF de detalhamento gerado: {nome_arquivo_pdf}")
        except KeyboardInterrupt:
            # Não propagar interrupção do usuário; registrar e seguir
            self._log_validacao('AVISO', 'Geração de PDF interrompida pelo usuário (KeyboardInterrupt). PDF não gerado.', {})
            _info("Geração de PDF interrompida pelo usuário. PDF não gerado.")
        except BaseException as e:
            # Captura falhas do reportlab (layout, encoding, etc.) e registra sem interromper o fluxo
            self._log_validacao('AVISO', f'Falha ao gerar PDF de detalhamento: {e}', {})
            _info(f"Aviso: falha ao gerar PDF de detalhamento: {e}")


    def _gerar_saida(self):
        """Gera o arquivo Excel com todas as abas de resultado e o PDF de detalhamento."""
        if not hasattr(self, 'comissoes_df') or self.comissoes_df.empty:
            _info("Nenhuma comissão foi calculada. O arquivo de saída não será gerado.")
            self._log_validacao("ERRO", "Cálculo final vazio", "Nenhuma comissão pôde ser calculada com base nos dados.")
            self.comissoes_df = pd.DataFrame()

        df_comissoes = self.comissoes_df
        # Reordenar colunas para melhor visualização
        if not df_comissoes.empty:
            colunas_principais = ['id_colaborador', 'nome_colaborador', 'cargo', 'processo', 'cod_produto', 'descricao_produto']
            colunas_contexto = ['linha', 'grupo', 'subgrupo', 'tipo_mercadoria']
            colunas_calculo = ['faturamento_item', 'taxa_rateio_aplicada', 'percentual_elegibilidade_pe', 'fator_correcao_fc']
            # detalhes do FC — para cada componente incluímos colunas padronizadas
            detalhes_shorts = ['fat_linha','conv_linha','fat_ind','conv_ind','rentab','retencao','forn1','forn2']
            detalhes_cols = []
            for s in detalhes_shorts:
                detalhes_cols.extend([
                    f'peso_{s}', f'realizado_{s}', f'meta_{s}', f'ating_{s}', f'ating_cap_{s}', f'comp_fc_{s}'
                ])
                if s.startswith('forn'):
                    detalhes_cols.append(f'moeda_{s}')

            colunas_resultado = ['comissao_potencial_maxima', 'comissao_calculada']
            ordem_final = colunas_principais + colunas_contexto + colunas_calculo + detalhes_cols + colunas_resultado
            # filtrar somente colunas que existem no DataFrame para evitar KeyError
            ordem_final = [c for c in ordem_final if c in df_comissoes.columns]
            df_comissoes = df_comissoes[ordem_final]

        # Remover do arquivo principal quaisquer colaboradores que recebem por recebimento.
        # Construir sets a partir de COMISSOES_RECEBIMENTO (quando presente) e também
        # incluir sempre os colaboradores detectados em self.recebe_por_recebimento.
        try:
            ids_to_remove = set()
            nomes_to_remove = set()

            if hasattr(self, 'comissoes_recebimento_df') and self.comissoes_recebimento_df is not None and not self.comissoes_recebimento_df.empty:
                df_rec = self.comissoes_recebimento_df
                for _, r in df_rec.iterrows():
                    cid = r.get('id_colaborador') if 'id_colaborador' in r.index else None
                    nome = r.get('nome_colaborador') if 'nome_colaborador' in r.index else None
                    if pd.notna(cid) and str(cid).strip() != '':
                        ids_to_remove.add(str(cid).strip())
                    if pd.notna(nome) and str(nome).strip() != '':
                        nomes_to_remove.add(str(nome).strip().lower())

            # Incluir colaboradores identificados para receber por recebimento, mesmo que
            # não existam linhas em COMISSOES_RECEBIMENTO (regra de negócio).
            try:
                for nome in getattr(self, 'recebe_por_recebimento', set()):
                    if not nome:
                        continue
                    # tentar achar id_colaborador correspondente
                    try:
                        df_colabs_com_cargos = self.data.get('COLABORADORES', pd.DataFrame())
                        row_col = df_colabs_com_cargos[df_colabs_com_cargos['nome_colaborador'].astype(str).str.strip().str.lower() == str(nome).strip().lower()]
                        if not row_col.empty:
                            cid = row_col.iloc[0].get('id_colaborador')
                            if pd.notna(cid) and str(cid).strip() != '':
                                ids_to_remove.add(str(cid).strip())
                            else:
                                nomes_to_remove.add(str(nome).strip().lower())
                        else:
                            nomes_to_remove.add(str(nome).strip().lower())
                    except Exception:
                        nomes_to_remove.add(str(nome).strip().lower())
            except Exception:
                pass

            if not df_comissoes.empty and (ids_to_remove or nomes_to_remove):
                before = len(df_comissoes)

                def _is_removed(row):
                    try:
                        cid = row.get('id_colaborador')
                        nome = row.get('nome_colaborador')
                        if pd.notna(cid) and str(cid).strip() in ids_to_remove:
                            return True
                        if pd.notna(nome) and str(nome).strip().lower() in nomes_to_remove:
                            return True
                        return False
                    except Exception:
                        return False

                df_comissoes = df_comissoes[~df_comissoes.apply(_is_removed, axis=1)].reset_index(drop=True)
                after = len(df_comissoes)
                removed_count = before - after
                removed_ids = sorted(list(ids_to_remove))
                removed_names = sorted(list(nomes_to_remove))
                if getattr(self, '_logger', None):
                    self._logger.info(f"Removidas {removed_count} linhas de COMISSOES_CALCULADAS para colaboradores que recebem por recebimento. ids_removidos={removed_ids} nomes_removidos={removed_names}")
        except Exception as e:
            self._log_validacao('AVISO', f'Falha ao filtrar COMISSOES_RECEBIMENTO do arquivo principal: {e}', {})


        # Construir resumo que inclui comissões por faturamento e por recebimento.
        # Objetivo: mostrar todos os colaboradores mesmo que recebam apenas por recebimento.
        resumo_cols = ['id_colaborador', 'nome_colaborador', 'cargo', 'comissao_total']
        df_resumo = pd.DataFrame(columns=resumo_cols)

        # Sumário das comissões calculadas (faturamento)
        if not df_comissoes.empty:
            res_fat = df_comissoes.groupby(['id_colaborador', 'nome_colaborador', 'cargo'])['comissao_calculada'].sum().reset_index()
            res_fat = res_fat.rename(columns={'comissao_calculada': 'comissao_total'})
        else:
            res_fat = pd.DataFrame(columns=resumo_cols)

        # Sumário das comissões por recebimento (se existir)
        res_rec = pd.DataFrame(columns=resumo_cols)
        if hasattr(self, 'comissoes_recebimento_df') and self.comissoes_recebimento_df is not None and not self.comissoes_recebimento_df.empty:
            try:
                res_rec = self.comissoes_recebimento_df.groupby(['id_colaborador', 'nome_colaborador', 'cargo'])['comissao_calculada'].sum().reset_index()
                res_rec = res_rec.rename(columns={'comissao_calculada': 'comissao_total'})
            except Exception:
                # Em caso de diferenças de colunas, tentar agrupar por nome apenas
                if 'nome_colaborador' in self.comissoes_recebimento_df.columns:
                    res_rec = self.comissoes_recebimento_df.groupby(['nome_colaborador'])['comissao_calculada'].sum().reset_index()
                    res_rec['id_colaborador'] = ''
                    res_rec['cargo'] = ''
                    res_rec = res_rec.rename(columns={'comissao_calculada': 'comissao_total'})

        # Concatenar e agregar por id/nome/cargo
        if not res_fat.empty or not res_rec.empty:
            df_resumo = pd.concat([res_fat, res_rec], ignore_index=True, sort=False).fillna(0)
            df_resumo = df_resumo.groupby(['id_colaborador', 'nome_colaborador', 'cargo'], dropna=False)['comissao_total'].sum().reset_index()
        else:
            df_resumo = pd.DataFrame(columns=resumo_cols)

        # Garantir que todos os colaboradores definidos em Regras (COLABORADORES) apareçam no resumo
        try:
            colabs = self.data.get('COLABORADORES') if isinstance(self.data, dict) else None
            if colabs is not None and not colabs.empty:
                # Normalizar nomes dos campos esperados
                id_field = None
                name_field = None
                cargo_field = None
                for c in colabs.columns:
                    lc = c.strip().lower()
                    if lc in ('id_colaborador', 'id', 'codigo', 'codigo_colaborador'):
                        id_field = c
                    if lc in ('nome_colaborador', 'nome', 'colaborador'):
                        name_field = c
                    if lc in ('cargo', 'role', 'função', 'funcao'):
                        cargo_field = c

                base_all = pd.DataFrame(columns=['id_colaborador', 'nome_colaborador', 'cargo'])
                if name_field:
                    base_all['nome_colaborador'] = colabs[name_field].astype(str).str.strip()
                else:
                    base_all['nome_colaborador'] = ''
                if id_field:
                    base_all['id_colaborador'] = colabs[id_field].astype(str).str.strip()
                else:
                    base_all['id_colaborador'] = ''
                if cargo_field:
                    base_all['cargo'] = colabs[cargo_field].astype(str).str.strip()
                else:
                    base_all['cargo'] = ''

                # Juntar com os valores calculados, garantindo zeros quando ausentes
                df_resumo = pd.merge(base_all.drop_duplicates(subset=['id_colaborador', 'nome_colaborador']), df_resumo, on=['id_colaborador', 'nome_colaborador'], how='left')
                # garantir que combine_first receba Series; criar série vazia quando cargo_y ausente
                if 'cargo_x' in df_resumo.columns:
                    if 'cargo_y' in df_resumo.columns:
                        df_resumo['cargo'] = df_resumo['cargo_x'].combine_first(df_resumo['cargo_y'])
                    else:
                        df_resumo['cargo'] = df_resumo['cargo_x'].fillna('')
                else:
                    df_resumo['cargo'] = df_resumo.get('cargo', '')
                # Normalizar colunas finais
                if 'comissao_total' not in df_resumo.columns:
                    df_resumo['comissao_total'] = df_resumo.get('comissao_total', 0.0)
                df_resumo = df_resumo[['id_colaborador', 'nome_colaborador', 'cargo', 'comissao_total']]
                df_resumo['comissao_total'] = pd.to_numeric(df_resumo['comissao_total'], errors='coerce').fillna(0.0)
        except Exception:
            # Se algo falhar, manter o resumo previamente calculado
            pass

        # detalhes do FC foram incorporados em self.comissoes_df
        df_validacao = pd.DataFrame(self.validation_log)
        df_debug_fornecedores = pd.DataFrame(self.debug_fornecedores)

        with pd.ExcelWriter(NOME_ARQUIVO_SAIDA, engine='openpyxl') as writer:
            # Se sinalizado que não houve faturamento no mês, inserir uma linha de aviso
            try:
                if getattr(self, '_no_faturamento_mes', False):
                    aviso_row = {
                        'id_colaborador': '',
                        'nome_colaborador': 'AVISO: Nenhum item faturado neste mês',
                        'cargo': '',
                        'processo': '',
                        'cod_produto': '',
                        'descricao_produto': '',
                        'linha': '', 'grupo': '', 'subgrupo': '', 'tipo_mercadoria': '',
                        'faturamento_item': 0.0, 'taxa_rateio_aplicada': 0.0, 'percentual_elegibilidade_pe': 0.0, 'fator_correcao_fc': 0.0,
                        'comissao_potencial_maxima': 0.0, 'comissao_calculada': 0.0
                    }
                    # Garantir que as colunas existam e inserir linha no topo
                    cols = df_comissoes.columns.tolist() if not df_comissoes.empty else list(aviso_row.keys())
                    df_aviso = pd.DataFrame([aviso_row], columns=cols)
                    if df_comissoes.empty:
                        df_to_write = df_aviso
                    else:
                        df_to_write = pd.concat([df_aviso, df_comissoes], ignore_index=True, sort=False)
                    df_to_write.to_excel(writer, sheet_name='COMISSOES_CALCULADAS', index=False)
                else:
                    df_comissoes.to_excel(writer, sheet_name='COMISSOES_CALCULADAS', index=False)
            except Exception:
                df_comissoes.to_excel(writer, sheet_name='COMISSOES_CALCULADAS', index=False)
            df_resumo.to_excel(writer, sheet_name='RESUMO_COLABORADOR', index=False)
            # Aba: COMISSOES_RECEBIMENTO (uma linha por pagamento do processo)
            try:
                if hasattr(self, 'comissoes_recebimento_df') and self.comissoes_recebimento_df is not None and not self.comissoes_recebimento_df.empty:
                    df_rec_raw = self.data.get('RECEBIMENTOS', pd.DataFrame())
                    base = self.comissoes_recebimento_df.copy()
                    # Validar que cada processo pertence a uma única linha; se mais de uma, avisar e excluir
                    try:
                        linhas_por_proc = base.groupby('processo')['linha'].nunique(dropna=False)
                        procs_multilinha = set(linhas_por_proc[linhas_por_proc > 1].index.tolist())
                        for p in procs_multilinha:
                            self._log_validacao('AVISO', f'Processo com múltiplas linhas na COMISSOES_RECEBIMENTO (ignorado): {p}', {'processo': p})
                        if procs_multilinha:
                            base = base[~base['processo'].isin(procs_multilinha)]
                    except Exception:
                        pass

                    # Derivar a linha do processo (única) para exibir
                    try:
                        linha_proc_map = base.groupby('processo')['linha'].agg(lambda s: s.dropna().iloc[0] if len(s.dropna())>0 else None).to_dict()
                        base['linha_processo'] = base['processo'].map(linha_proc_map)
                    except Exception:
                        base['linha_processo'] = None

                    # Anexar DATA_RECEBIMENTO por (processo, valor, sequencia) para diferenciar pagamentos repetidos
                    if df_rec_raw is not None and not df_rec_raw.empty and all(c in df_rec_raw.columns for c in ['PROCESSO','DATA_RECEBIMENTO','VALOR_RECEBIDO']):
                        df_r = df_rec_raw[['PROCESSO','DATA_RECEBIMENTO','VALOR_RECEBIDO']].copy()
                        df_r['PROCESSO'] = df_r['PROCESSO'].astype(str).str.strip()
                        df_r['_seq'] = df_r.groupby(['PROCESSO','VALOR_RECEBIDO']).cumcount()
                        base['processo'] = base['processo'].astype(str).str.strip()
                        base['_seq'] = base.groupby(['processo','faturamento_item']).cumcount()
                        df_join = base.merge(df_r, left_on=['processo','faturamento_item','_seq'], right_on=['PROCESSO','VALOR_RECEBIDO','_seq'], how='left')
                        for cdrop in ['PROCESSO','VALOR_RECEBIDO','_seq']:
                            if cdrop in df_join.columns:
                                df_join = df_join.drop(columns=[cdrop])
                    else:
                        df_join = base.copy()

                    # Agregar por processo + DATA + linha + colaborador (uma linha por pagamento/colaborador);
                    # se houver divergência de taxa/PE entre itens desse pagamento, detalhar por contexto e sinalizar.
                    gb_main = [k for k in ['processo','DATA_RECEBIMENTO','linha_processo','nome_colaborador'] if k in df_join.columns]
                    out_rows = []
                    for keys, g in df_join.groupby(gb_main, dropna=False):
                        try:
                            div_rateio = g['taxa_rateio_aplicada'].nunique(dropna=False) > 1 if 'taxa_rateio_aplicada' in g.columns else False
                            div_pe = g['percentual_elegibilidade_pe'].nunique(dropna=False) > 1 if 'percentual_elegibilidade_pe' in g.columns else False
                            if not (div_rateio or div_pe):
                                row = {
                                    'processo': keys[gb_main.index('processo')] if 'processo' in gb_main else None,
                                    'DATA_RECEBIMENTO': keys[gb_main.index('DATA_RECEBIMENTO')] if 'DATA_RECEBIMENTO' in gb_main else None,
                                    'linha': keys[gb_main.index('linha_processo')] if 'linha_processo' in gb_main else None,
                                    'nome_colaborador': keys[gb_main.index('nome_colaborador')] if 'nome_colaborador' in gb_main else None,
                                    'valor_recebido_total': g['faturamento_item'].sum() if 'faturamento_item' in g.columns else 0.0,
                                    'comissao_total': g['comissao_calculada'].sum() if 'comissao_calculada' in g.columns else 0.0,
                                    'taxa_rateio_aplicada': (g['taxa_rateio_aplicada'].iloc[0] if 'taxa_rateio_aplicada' in g.columns and len(g)>0 else None),
                                    'percentual_elegibilidade_pe': (g['percentual_elegibilidade_pe'].iloc[0] if 'percentual_elegibilidade_pe' in g.columns and len(g)>0 else None)
                                }
                                # percentual de comissão para referência
                                try:
                                    row['percentual_comissao'] = float(row['taxa_rateio_aplicada']) * float(row['percentual_elegibilidade_pe'])
                                except Exception:
                                    row['percentual_comissao'] = None
                                out_rows.append(row)
                            else:
                                # detalhar por contexto (linha/grupo/subgrupo/tipo) e sinalizar divergência
                                inner_keys = [c for c in ['linha','grupo','subgrupo','tipo_mercadoria'] if c in g.columns]
                                for _, gi in g.groupby(inner_keys, dropna=False):
                                    row = {
                                        'processo': keys[gb_main.index('processo')] if 'processo' in gb_main else None,
                                        'DATA_RECEBIMENTO': keys[gb_main.index('DATA_RECEBIMENTO')] if 'DATA_RECEBIMENTO' in gb_main else None,
                                        'linha': (gi['linha'].iloc[0] if 'linha' in gi.columns and len(gi)>0 else (keys[gb_main.index('linha_processo')] if 'linha_processo' in gb_main else None)),
                                        'nome_colaborador': keys[gb_main.index('nome_colaborador')] if 'nome_colaborador' in gb_main else None,
                                        'valor_recebido_total': gi['faturamento_item'].sum() if 'faturamento_item' in gi.columns else 0.0,
                                        'comissao_total': gi['comissao_calculada'].sum() if 'comissao_calculada' in gi.columns else 0.0,
                                        'taxa_rateio_aplicada': (gi['taxa_rateio_aplicada'].iloc[0] if 'taxa_rateio_aplicada' in gi.columns and len(gi)>0 else None),
                                        'percentual_elegibilidade_pe': (gi['percentual_elegibilidade_pe'].iloc[0] if 'percentual_elegibilidade_pe' in gi.columns and len(gi)>0 else None),
                                        'aviso_divergencia': 'Divergencia de taxa_rateio e/ou PE entre itens deste pagamento'
                                    }
                                    try:
                                        row['percentual_comissao'] = float(row['taxa_rateio_aplicada']) * float(row['percentual_elegibilidade_pe'])
                                    except Exception:
                                        row['percentual_comissao'] = None
                                    # incluir contexto somente quando divergente
                                    for c in ['grupo','subgrupo','tipo_mercadoria']:
                                        if c in gi.columns:
                                            row[c] = gi[c].iloc[0]
                                    out_rows.append(row)
                        except Exception:
                            continue
                    df_out = pd.DataFrame(out_rows)
                    # Garantir presença da coluna de comissão
                    try:
                        if not df_out.empty and 'comissao_total' not in df_out.columns:
                            df_out['comissao_total'] = 0.0
                    except Exception:
                        pass


                    # Ordenação
                    sort_cols = [c for c in ['processo','DATA_RECEBIMENTO'] if c in df_out.columns]
                    if sort_cols:
                        df_out = df_out.sort_values(sort_cols)

                    df_out.to_excel(writer, sheet_name='COMISSOES_RECEBIMENTO', index=False)
                    if getattr(self, '_logger', None):
                        self._logger.info('Aba COMISSOES_RECEBIMENTO escrita (1 linha por pagamento; inclui linha do processo).')
            except Exception as e:
                self._log_validacao('AVISO', f'Falha ao escrever COMISSOES_RECEBIMENTO: {e}', {})
            df_validacao.to_excel(writer, sheet_name='VALIDACAO', index=False)
            # Abas de DEBUG adicionais para diagnosticar COMISSOES_RECEBIMENTO
            try:
                rec_raw = self.data.get('RECEBIMENTOS', pd.DataFrame())
                if rec_raw is not None and not rec_raw.empty:
                    rec_raw.to_excel(writer, sheet_name='DEBUG_RECEBIMENTOS_RAW', index=False)
            except Exception as e:
                self._log_validacao('AVISO', f'Falha ao escrever DEBUG_RECEBIMENTOS_RAW: {e}', {})

            try:
                dbg_rec_df = pd.DataFrame(getattr(self, 'debug_recebimentos', []))
                if not dbg_rec_df.empty:
                    dbg_rec_df.to_excel(writer, sheet_name='DEBUG_RECEBIMENTOS', index=False)
            except Exception as e:
                self._log_validacao('AVISO', f'Falha ao escrever DEBUG_RECEBIMENTOS: {e}', {})

            try:
                env = []
                try:
                    rec = self.data.get('RECEBIMENTOS', pd.DataFrame())
                    env.append({'categoria':'RECEBIMENTOS_cols','detalhe': ', '.join([str(c) for c in rec.columns]) if rec is not None and not rec.empty else ''})
                except Exception:
                    pass
                try:
                    anal = self.data.get('ANALISE_COMERCIAL_COMPLETA', pd.DataFrame())
                    env.append({'categoria':'ANALISE_cols','detalhe': ', '.join([str(c) for c in anal.columns]) if anal is not None and not anal.empty else ''})
                except Exception:
                    pass
                try:
                    rcv_set = sorted(list(getattr(self, 'recebe_por_recebimento', set())))
                    env.append({'categoria':'recebe_por_recebimento','detalhe': ', '.join(rcv_set)})
                except Exception:
                    pass
                # merge com qualquer coleta anterior
                try:
                    env.extend(getattr(self, 'debug_env', []))
                except Exception:
                    pass
                dbg_env_df = pd.DataFrame(env)
                if not dbg_env_df.empty:
                    dbg_env_df.to_excel(writer, sheet_name='DEBUG_ENV', index=False)
            except Exception as e:
                self._log_validacao('AVISO', f'Falha ao escrever DEBUG_ENV: {e}', {})

            # Aba com amostra e colunas do ANALISE_COMERCIAL_COMPLETA para verificar headers
            try:
                anal = self.data.get('ANALISE_COMERCIAL_COMPLETA', pd.DataFrame())
                if anal is not None and not anal.empty:
                    info_rows = []
                    try:
                        info_rows.append({'colunas_analise': ', '.join([str(c) for c in anal.columns.tolist()])})
                    except Exception:
                        pass
                    df_info = pd.DataFrame(info_rows)
                    df_info.to_excel(writer, sheet_name='DEBUG_ANALISE_INFO', index=False)
                    # Amostra de colunas relevantes, quando existirem
                    wanted = [
                        'Processo','Negcio','Negocio','Grupo','Subgrupo','Tipo de Mercadoria','Tipo Mercadoria',
                        'Consultor Interno','Consultor','Representante-pedido','Representante'
                    ]
                    have = [c for c in wanted if c in anal.columns]
                    if have:
                        anal[have].head(200).to_excel(writer, sheet_name='DEBUG_ANALISE_SAMPLE', index=False)
            except Exception as e:
                self._log_validacao('AVISO', f'Falha ao escrever DEBUG_ANALISE_*: {e}', {})
            # Aba de depuração detalhada para metas de fornecedores
            # Limpeza persistente da aba DEBUG_FORNECEDORES: manter apenas uma linha por (colaborador, linha_item, fornecedor_index)
            try:
                if not df_debug_fornecedores.empty:
                    # Normalizar campos esperados
                    if 'peso_fornecedor' in df_debug_fornecedores.columns:
                        df_debug_fornecedores['peso_fornecedor'] = pd.to_numeric(df_debug_fornecedores['peso_fornecedor'], errors='coerce').fillna(0.0)

                    # Mapear atribuições por colaborador para identificar quais linhas devem ser consideradas
                    df_atr = self.data.get('ATRIBUICOES', pd.DataFrame())
                    atrib_map = {}
                    if not df_atr.empty and 'colaborador' in df_atr.columns and 'linha' in df_atr.columns:
                        for col, grp in df_atr.groupby('colaborador'):
                            atrib_map[col] = set(grp['linha'].dropna().tolist())

                    # Filtrar: manter apenas linhas onde o colaborador tem a linha atribuída e peso_fornecedor > 0
                    def _row_deve_ser_mantida(row):
                        try:
                            colab = row.get('colaborador')
                            linha = row.get('linha_item')
                            peso = float(row.get('peso_fornecedor') or 0.0)
                            if peso <= 0:
                                return False
                            if colab in atrib_map:
                                return (linha in atrib_map[colab])
                            # Se colaborador não tem atribuições registradas, assumir conservador: não manter
                            return False
                        except Exception:
                            return False

                    mask = df_debug_fornecedores.apply(_row_deve_ser_mantida, axis=1)
                    df_debug_fornecedores = df_debug_fornecedores[mask].copy()

                    # Deduplicar: manter apenas uma ocorrência por (colaborador, linha_item, fornecedor_index)
                    dedup_dbg = [c for c in ['colaborador', 'linha_item', 'fornecedor_index'] if c in df_debug_fornecedores.columns]
                    if dedup_dbg:
                        df_debug_fornecedores = df_debug_fornecedores.drop_duplicates(subset=dedup_dbg, keep='last').reset_index(drop=True)

                    # Escrever a aba de depuração limpa
                    if not df_debug_fornecedores.empty:
                        df_debug_fornecedores.to_excel(writer, sheet_name='DEBUG_FORNECEDORES', index=False)
            except Exception as e:
                self._log_validacao('AVISO', f'Falha ao limpar DEBUG_FORNECEDORES antes de salvar: {e}', {})

            # Aba: CROSS_SELLING_DECISIONS (registro das decisões tomadas)
            try:
                if self.cross_selling_decisions:
                    df_cs = pd.DataFrame([{
                        'processo': p,
                        'consultor': v.get('consultor'),
                        'linha': v.get('linha'),
                        'taxa_pct': v.get('taxa'),
                        'decision': v.get('decision'),
                        'timestamp': v.get('timestamp')
                    } for p, v in self.cross_selling_decisions.items()])
                    df_cs.to_excel(writer, sheet_name='CROSS_SELLING_DECISIONS', index=False)
            except Exception as e:
                self._log_validacao('AVISO', f'Falha ao escrever CROSS_SELLING_DECISIONS: {e}', {})
            # Aba: RECONCILIACAO (detalhamento e resumo de saldos)
            try:
                detalhes = getattr(self, 'reconciliacao_detalhada_list', [])
                resumo = getattr(self, 'reconciliacao_resumo_list', [])

                if detalhes:
                    df_reconciliacao_detalhada = pd.DataFrame(detalhes)
                    if hasattr(self, 'comissoes_df') and isinstance(self.comissoes_df, pd.DataFrame) and not self.comissoes_df.empty:
                        colunas_base = [c for c in self.comissoes_df.columns if c in df_reconciliacao_detalhada.columns]
                        colunas_extra = [c for c in df_reconciliacao_detalhada.columns if c not in colunas_base]
                        if colunas_base:
                            df_reconciliacao_detalhada = df_reconciliacao_detalhada[colunas_base + colunas_extra]
                else:
                    _info("Nenhuma reconciliação foi processada. Aba 'RECONCILIACAO' ficará vazia.")
                    if hasattr(self, 'comissoes_df') and isinstance(self.comissoes_df, pd.DataFrame):
                        df_reconciliacao_detalhada = pd.DataFrame(columns=self.comissoes_df.columns)
                    else:
                        df_reconciliacao_detalhada = pd.DataFrame()

                df_reconciliacao_detalhada.to_excel(writer, sheet_name='RECONCILIACAO', index=False)
                _info(f"Aba 'RECONCILIACAO' (Detalhada) gerada com {len(df_reconciliacao_detalhada)} linhas.")

                if resumo:
                    df_reconciliacao_resumo = pd.DataFrame(resumo)
                    start_row_resumo = len(df_reconciliacao_detalhada) + 3
                    df_reconciliacao_resumo.to_excel(
                        writer,
                        sheet_name='RECONCILIACAO',
                        index=False,
                        startrow=start_row_resumo
                    )
                    _info("Resumo de reconciliação adicionado ao final da aba.")
            except Exception as e:
                self._log_validacao('AVISO', f'Falha ao escrever RECONCILIACAO: {e}', {})

            # Aba: ESTADO (salva o snapshot atual do estado de recebimentos/reconciliações)
            try:
                if getattr(self, 'estado', None) is not None and not self.estado.empty:
                    self.estado.to_excel(writer, sheet_name='ESTADO', index=False)
            except Exception as e:
                self._log_validacao('AVISO', f'Falha ao escrever ESTADO no arquivo de saída: {e}', {})
        
        try:
            style_output_workbook(NOME_ARQUIVO_SAIDA)
        except Exception:
            pass

        _info(f"\nCálculo finalizado. Arquivo de saída Excel gerado: {NOME_ARQUIVO_SAIDA}")
        _info("\n--- RESUMO POR COLABORADOR ---")
        _info(df_resumo.to_string(index=False))

        try:
            # Logamos aqui a tentativa explícita de gerar o PDF para facilitar diagnóstico
            if REPORTLAB_DISPONIVEL:
                if getattr(self, '_logger', None):
                    self._logger.info('Tentativa de gerar PDF de detalhamento (reportlab disponível)')
                self._gerar_detalhamento_pdf()
            else:
                _info('\nAVISO: A biblioteca "reportlab" não está instalada. O PDF não será gerado.')
                self._log_validacao('AVISO', 'Biblioteca reportlab ausente; PDF não gerado.', {})
        except Exception as e:
            if not REPORTLAB_DISPONIVEL:
                _info("\nAVISO: A biblioteca 'reportlab' não está instalada.")
                _info("Para gerar o PDF de detalhamento, instale-a com: pip install reportlab")
            else:
                _info(f"\nOcorreu um erro ao gerar o PDF de detalhamento: {e}")

    def executar(self):
        """Executa o fluxo completo de cálculo de comissões."""
        _info("Iniciando cálculo de comissões...")
        _phase("1. Carregando arquivos...")
        self._carregar_dados()
        _phase("2. Validando dados...")
        self._validar_dados()
        _phase("3. Pré-processando informações...")
        self._preprocessar_dados()
        _phase("4. Calculando valores realizados agregados...")
        self._calcular_realizado()
        _phase("5. Calculando comissões e FC item a item...")
        self._calcular_comissoes()
        # Carregar estado (se existir) e processar recebimentos/reconciliações
        _phase("5.1 Carregando estado de recebimentos e aplicando adiantamentos (se houver)...")
        self._carregar_estado()
        self._aplicar_adiantamentos_recebimentos()
        _phase("5.2 Executando reconciliações de processos quitados...")
        self._executar_reconciliacoes()
        _phase("6. Gerando arquivos de saída...")
        self._gerar_saida()
        # Salvar estado persistente (obrigatório)
        try:
            self._salvar_estado()
        except Exception:
            pass


if __name__ == '__main__':
    try:
        # Suporte a modo não-interativo via CLI e variáveis de ambiente
        def _parse_cli_env_mes_ano():
            mes_cli = None
            ano_cli = None
            skip_clean = False
            try:
                import argparse
                parser = argparse.ArgumentParser(add_help=False)
                parser.add_argument('--mes', type=int)
                parser.add_argument('--ano', type=int)
                parser.add_argument('--skip-clean', action='store_true', default=False)
                args, _ = parser.parse_known_args()
                mes_cli = args.mes
                ano_cli = args.ano
                skip_clean = bool(args.skip_clean)
            except Exception:
                pass

            # Variáveis de ambiente alternativas aceitas
            env_mes = os.environ.get('MES_APURACAO') or os.environ.get('COMISSOES_MES') or os.environ.get('MES')
            env_ano = os.environ.get('ANO_APURACAO') or os.environ.get('COMISSOES_ANO') or os.environ.get('ANO')
            if mes_cli is None and env_mes:
                try:
                    mes_cli = int(str(env_mes).strip())
                except Exception:
                    mes_cli = None
            if ano_cli is None and env_ano:
                try:
                    ano_cli = int(str(env_ano).strip())
                except Exception:
                    ano_cli = None

            # Flag para pular limpeza por ENV
            if not skip_clean:
                skip_clean = os.environ.get('SKIP_CLEAN') in ('1', 'true', 'TRUE', 'yes', 'YES')

            return mes_cli, ano_cli, skip_clean

        # Perguntar somente se não vier por CLI/ENV
        def solicitar_mes_ano():
            mes_cli, ano_cli, skip_clean = _parse_cli_env_mes_ano()
            if isinstance(mes_cli, int) and 1 <= mes_cli <= 12 and isinstance(ano_cli, int) and 2000 < ano_cli < 2100:
                return mes_cli, ano_cli, skip_clean
            try:
                from preparar_dados_mensais import obter_mes_ano
                m, a = obter_mes_ano()
                return m, a, False
            except Exception:
                # fallback simples (interativo)
                while True:
                    try:
                        ano = int(input("Digite o ano para apuração (ex: 2025): "))
                        if 2000 < ano < 2100:
                            break
                    except Exception:
                        pass

                while True:
                    try:
                        mes = int(input(f"Digite o número do mês para apuração em {ano} (1-12): "))
                        if 1 <= mes <= 12:
                            break
                    except Exception:
                        pass

                return mes, ano, False

        mes, ano, skip_clean = solicitar_mes_ano()
        # Sempre executar os scripts de limpeza que geram Recebimentos_do_Mes.xlsx
        # e Status_Pagamentos_Processos.xlsx, e em seguida o preparador de dados
        # para garantir que os arquivos necessários sejam gerados para o mês/ano selecionado.
        try:
            import subprocess, os
            if skip_clean:
                _info(f"PULANDO scripts de limpeza por flag --skip-clean/ENV para {mes}/{ano}.")
            else:
                _info(f"Executando script de limpeza de recebimentos para {mes}/{ano}...")
                r1 = subprocess.run([sys.executable, 'limpeza_recebimentos.py', str(mes), str(ano)], text=True, check=False)
                if r1.returncode != 0:
                    print(f"ERRO: o script 'limpeza_recebimentos.py' retornou código {r1.returncode}. Abortando.")
                    sys.exit(1)
                if not os.path.exists('Recebimentos_do_Mes.xlsx'):
                    print("ERRO: arquivo 'Recebimentos_do_Mes.xlsx' não foi gerado pelo script de limpeza. Abortando.")
                    sys.exit(1)

                _info(f"Executando script de limpeza de status de pagamentos para {mes}/{ano}...")
                r2 = subprocess.run([sys.executable, 'limpeza_status_pagamentos.py', str(mes), str(ano)], text=True, check=False)
                if r2.returncode != 0:
                    print(f"ERRO: o script 'limpeza_status_pagamentos.py' retornou código {r2.returncode}. Abortando.")
                    sys.exit(1)
                if not os.path.exists('Status_Pagamentos_Processos.xlsx'):
                    print("ERRO: arquivo 'Status_Pagamentos_Processos.xlsx' não foi gerado pelo script de limpeza. Abortando.")
                    sys.exit(1)
                _info("Scripts de limpeza executados com sucesso.")
        except Exception as e:
            print(f"AVISO: falha ao executar os scripts de limpeza automaticamente: {e}. Abortando.")
            sys.exit(1)

    # Sempre executar o preparador de dados no início para garantir que os arquivos
    # Faturados.xlsx, Conversões.xlsx, Faturados_YTD.xlsx e Retencao_Clientes.xlsx
    # sejam gerados para o mês/ano selecionado.
        try:
            _info(f"Executando o preparador de dados para {mes}/{ano}...")
            import preparar_dados_mensais
            if not preparar_dados_mensais.run_preparador(mes, ano):
                print("ERRO: O script 'preparar_dados_mensais.py' encontrou um erro. Abortando.")
                sys.exit(1)
            _info("Preparador de dados executado com sucesso.")
        except Exception as e:
            print(f"AVISO: falha ao executar o preparador de dados automaticamente: {e}. Abortando.")
            sys.exit(1)
        # Atualizar variáveis de arquivo para usar arquivos gerados pelo preparador (os nomes fixos esperados)
        ARQUIVO_FATURADOS = "Faturados.xlsx"
        ARQUIVO_CONVERSOES = "Conversões.xlsx"
        ARQUIVO_FATURADOS_YTD = "Faturados_YTD.xlsx"

        # Selecionar o arquivo de rentabilidade agrupada correto na pasta 'rentabilidades'
        mm = str(mes).zfill(2)
        candidato = f"rentabilidades/rentabilidade_{mm}_{ano}_agrupada.xlsx"
        import glob
        encontrados = glob.glob(f"rentabilidades/*{mm}*{ano}*agrupada*.xlsx")
        if encontrados:
            ARQUIVO_RENTABILIDADE = encontrados[0]
            _info(f"Usando arquivo de rentabilidade: {ARQUIVO_RENTABILIDADE}")
        else:
            # fallback para nome padrão caso não encontre agrupada
            padrao = f"rentabilidades/rentabilidade_{mm}_{ano}_agrupada.xlsx"
            if os.path.exists(padrao):
                ARQUIVO_RENTABILIDADE = padrao
                _info(f"Usando arquivo de rentabilidade: {ARQUIVO_RENTABILIDADE}")
            else:
                _info(f"Aviso: não foi encontrado arquivo de rentabilidade agrupada para {mm}/{ano} na pasta 'rentabilidades'. Procurados: {candidato}")

        calculadora = CalculoComissao()
        calculadora.executar()
    except Exception as e:
        print(f"\nOcorreu um erro fatal durante a execução: {e}")
